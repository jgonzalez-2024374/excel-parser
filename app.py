from flask import Flask, request, jsonify, send_file, after_this_request
from flask import render_template
from flask import make_response
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import datetime, date
from functools import wraps
from difflib import SequenceMatcher
import os
import io
import tempfile
import unicodedata
import re
import json
import pickle
import shutil
import time
import hmac
import base64
import mimetypes
import urllib.request
import xml.etree.ElementTree as ET

app = Flask(__name__)


# ============================================================
# SEGURIDAD API / RUTAS CONFIDENCIALES
# ============================================================

API_KEY_ENV_NAME = "INTELFON_API_KEY"

def _api_key_valida():
    """Valida X-API-Key contra la variable segura configurada en Render."""
    expected = str(os.environ.get(API_KEY_ENV_NAME, "") or "").strip()
    provided = str(request.headers.get("X-API-Key", "") or "").strip()

    # Falla cerrado: si Render no tiene la variable, no expone datos bancarios.
    if not expected or not provided:
        return False

    return hmac.compare_digest(provided, expected)


def require_api_key(func):
    """Protege endpoints confidenciales sin alterar su lógica interna."""
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not _api_key_valida():
            response = jsonify({
                "success": False,
                "error": "No autorizado"
            })
            response.status_code = 401
            response.headers["Cache-Control"] = "no-store"
            return response

        return func(*args, **kwargs)

    return wrapper

# Identificador visible para confirmar qué versión está ejecutando Render.
APP_BUILD = "dashboard-v4.9-saldo-disponible-multimoneda-20260831"


# ============================================================
# CONFIGURACIÓN
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATES_DIR = os.path.join(
    BASE_DIR,
    "templates"
)

TEMPLATE_FILE = os.path.join(
    TEMPLATES_DIR,
    "Consolidación Estados Financieros e Ingresos - Plantilla.xlsx"
)

OUTPUT_FILENAME = "Consolidación Estados Financieros e Ingresos.xlsx"

SHEET_SALDOS = "SALDOS POR CUENTA"
SHEET_ESTADOS = "ESTADOS CONSOLIDADOS"
SHEET_REPORTE = "REPORTE CREDITOS DIARIOS"
SHEET_TABLERO = "TABLERO CREDITOS"


# ============================================================
# CONFIGURACIÓN DASHBOARD HTML
# ============================================================

DASHBOARD_TEMPLATE = os.path.join(
    TEMPLATES_DIR,
    "dashboard",
    "index.html"
)

DASHBOARD_DATA_DIR = os.path.join(
    BASE_DIR,
    "data"
)

DASHBOARD_DATA_FILE = os.path.join(
    DASHBOARD_DATA_DIR,
    "dashboard_data.json"
)

os.makedirs(
    DASHBOARD_DATA_DIR,
    exist_ok=True
)


# ============================================================
# PROCESAMIENTO POR LOTES
# ============================================================

# Los lotes viven en /tmp mientras dura la ejecución/instancia de Render.
# Es suficiente para el flujo de Make: todos los archivos se procesan
# consecutivamente y luego se finaliza el lote.
BATCH_ROOT = os.path.join(
    tempfile.gettempdir(),
    "excel_parser_batches"
)

os.makedirs(
    BATCH_ROOT,
    exist_ok=True
)

BATCH_MAX_AGE_SECONDS = 6 * 60 * 60


def _sanitizar_batch_id(value):
    """
    Permite únicamente caracteres seguros para usar el batch_id en rutas.
    """
    value = str(value or "").strip()

    if not value:
        return ""

    value = re.sub(
        r"[^A-Za-z0-9_-]",
        "_",
        value
    )

    return value[:120]


def _batch_dir(batch_id):
    batch_id = _sanitizar_batch_id(
        batch_id
    )

    if not batch_id:
        raise ValueError(
            "batch_id vacío o inválido"
        )

    return os.path.join(
        BATCH_ROOT,
        batch_id
    )


def _batch_state_path(batch_id):
    return os.path.join(
        _batch_dir(batch_id),
        "state.pkl"
    )


def _batch_result_path(batch_id):
    return os.path.join(
        _batch_dir(batch_id),
        "Consolidacion_Estados_Financieros.xlsx"
    )


def _limpiar_lotes_antiguos():
    """
    Evita acumular archivos temporales indefinidamente.
    """
    try:
        now_ts = time.time()

        for name in os.listdir(
            BATCH_ROOT
        ):
            path = os.path.join(
                BATCH_ROOT,
                name
            )

            if not os.path.isdir(
                path
            ):
                continue

            try:
                age = (
                    now_ts
                    - os.path.getmtime(
                        path
                    )
                )

                if age > BATCH_MAX_AGE_SECONDS:
                    shutil.rmtree(
                        path,
                        ignore_errors=True
                    )

            except Exception:
                pass

    except Exception:
        pass


def _crear_estado_lote(batch_id):
    now_iso = datetime.now().isoformat(
        timespec="seconds"
    )

    return {
        "batch_id": batch_id,
        "created_at": now_iso,
        "updated_at": now_iso,
        "finalized": False,
        "files": {},
        "dashboard_payload": None,
        "result_path": None
    }


def _cargar_estado_lote(
    batch_id,
    create=False
):
    batch_id = _sanitizar_batch_id(
        batch_id
    )

    if not batch_id:
        raise ValueError(
            "batch_id requerido"
        )

    directory = _batch_dir(
        batch_id
    )

    state_path = _batch_state_path(
        batch_id
    )

    if os.path.exists(
        state_path
    ):
        with open(
            state_path,
            "rb"
        ) as fh:
            state = pickle.load(
                fh
            )

        return state

    if not create:
        return None

    os.makedirs(
        directory,
        exist_ok=True
    )

    state = _crear_estado_lote(
        batch_id
    )

    _guardar_estado_lote(
        batch_id,
        state
    )

    return state


def _guardar_estado_lote(
    batch_id,
    state
):
    directory = _batch_dir(
        batch_id
    )

    os.makedirs(
        directory,
        exist_ok=True
    )

    state["updated_at"] = datetime.now().isoformat(
        timespec="seconds"
    )

    destination = _batch_state_path(
        batch_id
    )

    temp_path = (
        destination
        + ".tmp"
    )

    with open(
        temp_path,
        "wb"
    ) as fh:
        pickle.dump(
            state,
            fh,
            protocol=pickle.HIGHEST_PROTOCOL
        )

    os.replace(
        temp_path,
        destination
    )


def _agregar_archivo_a_lote(
    batch_id,
    file_id,
    file_name,
    transactions,
    processed_sheets,
    currency
):
    """
    Agrega/reemplaza un archivo dentro del mismo lote.

    file_id permite que un retry de Make no duplique movimientos.
    """
    _limpiar_lotes_antiguos()

    batch_id = _sanitizar_batch_id(
        batch_id
    )

    state = _cargar_estado_lote(
        batch_id,
        create=True
    )

    file_key = str(
        file_id
        or file_name
        or (
            "file_"
            + str(
                len(
                    state.get(
                        "files",
                        {}
                    )
                )
                + 1
            )
        )
    ).strip()

    state.setdefault(
        "files",
        {}
    )

    state["files"][
        file_key
    ] = {
        "file_id": str(
            file_id
            or ""
        ),
        "file_name": str(
            file_name
            or ""
        ),
        "currency": str(
            currency
            or "N/D"
        ),
        "processed_sheets": processed_sheets,
        "transactions": transactions,
        "received_at": datetime.now().isoformat(
            timespec="seconds"
        )
    }

    # Si se vuelve a usar el mismo batch después de finalizarlo,
    # obligar a finalizar nuevamente con el conjunto actualizado.
    state["finalized"] = False
    state["dashboard_payload"] = None
    state["result_path"] = None

    _guardar_estado_lote(
        batch_id,
        state
    )

    return state


def _transacciones_del_lote(
    state
):
    transactions = []

    for file_info in state.get(
        "files",
        {}
    ).values():

        transactions.extend(
            file_info.get(
                "transactions",
                []
            )
            or []
        )

    return transactions


def _moneda_preferida_lote(
    state
):
    currencies = {
        str(
            info.get(
                "currency",
                "N/D"
            )
        ).upper()
        for info in state.get(
            "files",
            {}
        ).values()
        if str(
            info.get(
                "currency",
                "N/D"
            )
        ).upper()
        not in {
            "",
            "N/D",
            "MULTI"
        }
    }

    if len(currencies) == 1:
        return next(
            iter(
                currencies
            )
        )

    if len(currencies) > 1:
        return "MULTI"

    return "N/D"


def _dashboard_lote(
    batch_id
):
    state = _cargar_estado_lote(
        batch_id,
        create=False
    )

    if not state:
        return None

    return state.get(
        "dashboard_payload"
    )


# Cache del tipo de cambio de referencia USD/GTQ.
# El Banco de Guatemala publica un valor diario; se refresca cada 15 minutos
# para evitar llamadas innecesarias al servicio externo.
FX_CACHE = {
    "rate_gtq_per_usd": None,
    "date": "",
    "fetched_at": 0.0,
    "source": "Banco de Guatemala"
}

FX_CACHE_TTL_SECONDS = 15 * 60

DASHBOARD_CACHE = {
    "meta": {
        "file": "",
        "fileCut": "",
        "latestMovement": "",
        "currency": "N/D",
        "accounts": 0,
        "banks": 0,
        "movementCount": 0
    },
    "totals": {
        "initial": 0.0,
        "final": 0.0,
        "change": 0.0,
        "changePct": 0.0,
        "credits": 0.0,
        "debits": 0.0,
        "netFlow": 0.0
    },
    "highlights": {
        "maxCredit": None,
        "maxDebit": None,
        "maxNet": None,
        "minNet": None
    },
    "banks": [],
    "accounts": [],
    "daily": [],
    "transactions": []
}


# Versión del generador del dashboard descargable.
DASHBOARD_ATTACHMENT_VERSION = "executive-dynamic-3.0"


# ============================================================
# INICIO
# ============================================================

@app.route("/", methods=["GET"])
def home():
    return jsonify({
        "status": "ok",
        "app_build": APP_BUILD,
        "message": "Excel Parser funcionando",
        "parser_version": PARSER_VERSION if "PARSER_VERSION" in globals() else "legacy",
        "template_exists": os.path.exists(TEMPLATE_FILE),
        "endpoint": "/process-excel",
        "batch_endpoints": {
            "finalize": "/finalize-batch",
            "excel": "/result-excel?batch_id=...",
            "dashboard": "/dashboard-file?batch_id=...",
            "status": "/batch-status?batch_id=..."
        }
    })




# ============================================================
# TIPO DE CAMBIO USD / GTQ
# ============================================================

BANGUAT_FX_URL = (
    "https://www.banguat.gob.gt/"
    "variables/ws/TipoCambio.asmx"
)


def _consultar_tipo_cambio_banguat():
    """
    Consulta TipoCambioDia del Web Service del Banco de Guatemala.

    Retorna:
        {
            "rate_gtq_per_usd": float,
            "date": str,
            "source": "Banco de Guatemala"
        }

    No utiliza librerías externas.
    """

    soap_body = """<?xml version="1.0" encoding="utf-8"?>
<soap:Envelope
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
    xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
  <soap:Body>
    <TipoCambioDia xmlns="http://www.banguat.gob.gt/variables/ws/" />
  </soap:Body>
</soap:Envelope>""".encode("utf-8")

    req = urllib.request.Request(
        BANGUAT_FX_URL,
        data=soap_body,
        method="POST",
        headers={
            "Content-Type": "text/xml; charset=utf-8",
            "SOAPAction": (
                '"http://www.banguat.gob.gt/'
                'variables/ws/TipoCambioDia"'
            ),
            "User-Agent": "excel-parser/1.0"
        }
    )

    with urllib.request.urlopen(
        req,
        timeout=10
    ) as response:

        xml_bytes = response.read()

    root = ET.fromstring(
        xml_bytes
    )

    referencia = None
    fecha = ""

    for element in root.iter():
        local_name = element.tag.split("}")[-1].lower()

        if (
            local_name == "referencia"
            and referencia is None
        ):
            try:
                referencia = float(
                    str(
                        element.text
                        or ""
                    ).strip()
                )
            except Exception:
                pass

        if (
            local_name == "fecha"
            and not fecha
        ):
            fecha = str(
                element.text
                or ""
            ).strip()

    if (
        referencia is None
        or referencia <= 0
    ):
        raise ValueError(
            "Banco de Guatemala no devolvió una referencia válida."
        )

    return {
        "rate_gtq_per_usd": referencia,
        "date": fecha,
        "source": "Banco de Guatemala"
    }


@app.route(
    "/api/tipo-cambio",
    methods=["GET"]
)
def tipo_cambio_actual():
    """
    Devuelve el tipo de cambio de referencia USD -> GTQ.

    Se permite CORS porque el dashboard descargado puede abrirse desde file://
    y consultar este endpoint alojado en Render.
    """

    force_refresh = (
        request.args.get(
            "refresh",
            ""
        )
        == "1"
    )

    now_ts = time.time()

    cache_rate = FX_CACHE.get(
        "rate_gtq_per_usd"
    )

    cache_age = (
        now_ts
        - float(
            FX_CACHE.get(
                "fetched_at",
                0.0
            )
            or 0.0
        )
    )

    if (
        not force_refresh
        and cache_rate
        and cache_age < FX_CACHE_TTL_SECONDS
    ):
        payload = {
            "success": True,
            "rate_gtq_per_usd": cache_rate,
            "date": FX_CACHE.get(
                "date",
                ""
            ),
            "source": FX_CACHE.get(
                "source",
                "Banco de Guatemala"
            ),
            "cached": True,
            "stale": False,
            "fetched_at": datetime.now().isoformat(
                timespec="seconds"
            )
        }

        response = jsonify(
            payload
        )

        response.headers[
            "Access-Control-Allow-Origin"
        ] = "*"

        response.headers[
            "Cache-Control"
        ] = "no-store"

        return response

    try:
        live = _consultar_tipo_cambio_banguat()

        FX_CACHE.update({
            "rate_gtq_per_usd": live[
                "rate_gtq_per_usd"
            ],
            "date": live.get(
                "date",
                ""
            ),
            "source": live.get(
                "source",
                "Banco de Guatemala"
            ),
            "fetched_at": now_ts
        })

        payload = {
            "success": True,
            "rate_gtq_per_usd": FX_CACHE[
                "rate_gtq_per_usd"
            ],
            "date": FX_CACHE[
                "date"
            ],
            "source": FX_CACHE[
                "source"
            ],
            "cached": False,
            "stale": False,
            "fetched_at": datetime.now().isoformat(
                timespec="seconds"
            )
        }

        status_code = 200

    except Exception as error:

        # Si hubo una consulta exitosa antes, conservar el último dato
        # como respaldo en lugar de dejar el dashboard inutilizable.
        if FX_CACHE.get(
            "rate_gtq_per_usd"
        ):
            payload = {
                "success": True,
                "rate_gtq_per_usd": FX_CACHE[
                    "rate_gtq_per_usd"
                ],
                "date": FX_CACHE.get(
                    "date",
                    ""
                ),
                "source": FX_CACHE.get(
                    "source",
                    "Banco de Guatemala"
                ),
                "cached": True,
                "stale": True,
                "warning": (
                    "No se pudo actualizar el tipo de cambio; "
                    "se muestra el último valor disponible."
                ),
                "fetched_at": datetime.now().isoformat(
                    timespec="seconds"
                )
            }

            status_code = 200

        else:
            payload = {
                "success": False,
                "error": str(
                    error
                ),
                "source": "Banco de Guatemala"
            }

            status_code = 502

    response = jsonify(
        payload
    )

    response.headers[
        "Access-Control-Allow-Origin"
    ] = "*"

    response.headers[
        "Cache-Control"
    ] = "no-store"

    return (
        response,
        status_code
    )


# ============================================================
# DASHBOARD HTML
# ============================================================

@app.route(
    "/dashboard",
    methods=["GET"],
    strict_slashes=False
)
@require_api_key
def dashboard():
    """Muestra la plantilla ejecutiva con los datos del último Excel procesado."""

    if not os.path.exists(DASHBOARD_TEMPLATE):
        return jsonify({
            "success": False,
            "error": "No existe index.html del dashboard",
            "path": DASHBOARD_TEMPLATE
        }), 404

    try:
        batch_id = _sanitizar_batch_id(
            request.args.get(
                "batch_id",
                ""
            )
        )

        datos = None

        if batch_id:
            datos = _dashboard_lote(
                batch_id
            )

            if datos is None:
                return jsonify({
                    "success": False,
                    "error": "Lote no encontrado o todavía no finalizado",
                    "batch_id": batch_id
                }), 404

        html = renderizar_dashboard_nueva_plantilla(
            datos
        )
        response = make_response(html)
        response.headers["Content-Type"] = "text/html; charset=utf-8"
        response.headers["Cache-Control"] = (
            "no-store, no-cache, must-revalidate, max-age=0"
        )
        return response

    except Exception as error:
        return jsonify({
            "success": False,
            "error": str(error)
        }), 500


@app.route(
    "/data.json",
    methods=["GET"]
)
@app.route(
    "/dashboard/data.json",
    methods=["GET"]
)
@require_api_key
def dashboard_data_endpoint():

    batch_id = _sanitizar_batch_id(
        request.args.get(
            "batch_id",
            ""
        )
    )

    data = (
        _dashboard_lote(
            batch_id
        )
        if batch_id
        else cargar_dashboard_data()
    )

    if data is None:
        return jsonify({
            "success": False,
            "error": "Lote no encontrado o todavía no finalizado",
            "batch_id": batch_id
        }), 404

    response = jsonify(
        data
    )

    response.headers["Cache-Control"] = (
        "no-store, no-cache, must-revalidate, max-age=0"
    )

    return response


@app.route(
    "/dashboard-file",
    methods=["GET"],
    strict_slashes=False
)
@require_api_key
def dashboard_file():
    """
    Descarga una copia autónoma del dashboard ejecutivo.
    Python sustituye BASE_DATA y RAW_TRANSACTIONS dentro del HTML,
    por lo que el archivo puede abrirse localmente sin data.json y sin CORS.
    """

    try:
        batch_id = _sanitizar_batch_id(
            request.args.get(
                "batch_id",
                ""
            )
        )

        datos = None

        if batch_id:
            datos = _dashboard_lote(
                batch_id
            )

            if datos is None:
                return jsonify({
                    "success": False,
                    "error": "Lote no encontrado o todavía no finalizado",
                    "batch_id": batch_id
                }), 404

        html = renderizar_dashboard_autonomo(
            datos
        )

        fecha_nombre = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        nombre_archivo = (
            "Dashboard_Saldos_Bancarios_"
            + fecha_nombre
            + ".html"
        )

        response = make_response(html)
        response.headers["Content-Type"] = "text/html; charset=utf-8"
        response.headers["Content-Disposition"] = (
            'attachment; filename="'
            + nombre_archivo
            + '"'
        )
        response.headers["Cache-Control"] = (
            "no-store, no-cache, must-revalidate, max-age=0"
        )
        response.headers["X-Dashboard-Attachment-Version"] = (
            DASHBOARD_ATTACHMENT_VERSION
        )
        response.headers["X-Dashboard-Filename"] = nombre_archivo

        return response

    except Exception as error:
        return jsonify({
            "success": False,
            "error": str(error)
        }), 500


# ============================================================
# UTILIDADES PARA CELDAS COMBINADAS
# ============================================================

def es_celda_combinada(cell):
    """
    Devuelve True si la celda es una MergedCell.
    """
    return isinstance(cell, MergedCell)


def escribir_celda_segura(worksheet, row, column, value):
    """
    Escribe en una celda solamente si NO es una MergedCell.

    Esto evita:
        'MergedCell' object attribute error
    """
    cell = worksheet.cell(
        row=row,
        column=column
    )

    if es_celda_combinada(cell):
        return False

    cell.value = value
    return True


def limpiar_celda_segura(worksheet, row, column):
    """
    Limpia una celda solamente si es editable.
    """
    cell = worksheet.cell(
        row=row,
        column=column
    )

    if es_celda_combinada(cell):
        return False

    cell.value = None
    return True


def obtener_celda_segura(worksheet, row, column):
    """
    Devuelve la celda si es editable.
    """
    cell = worksheet.cell(
        row=row,
        column=column
    )

    if es_celda_combinada(cell):
        return None

    return cell


def obtener_column_letter_seguro(column):
    """
    IMPORTANTE:
    No usar:

        worksheet.cell(...).column_letter

    porque puede devolver MergedCell.

    Usamos directamente get_column_letter().
    """
    return get_column_letter(column)


# ============================================================
# PARSER UNIVERSAL / NORMALIZACIÓN
# ============================================================

PARSER_VERSION = "universal-3.5-safe-currency-bank-logo-ocr"

# Este parser NO depende de un banco concreto. Las listas siguientes son
# vocabulario contable para reconocer columnas, no formatos rígidos por banco.
HEADER_ALIASES = {
    "fecha": [
        "fecha", "date", "fecha movimiento", "fecha de movimiento",
        "fecha transaccion", "fecha de transaccion", "transaction date",
        "fecha operacion", "fecha de operacion", "posting date", "value date",
    ],
    "referencia": [
        "referencia", "reference", "ref", "no doc", "no. doc", "nro doc",
        "documento", "numero documento", "numero de documento", "comprobante",
        "numero operacion", "operacion", "transaction id", "trace", "secuencial",
    ],
    "codigo": [
        "codigo", "code", "codigo transaccion", "codigo de transaccion",
        "transaction code", "tt", "tipo transaccion", "tipo de transaccion",
    ],
    "descripcion": [
        "descripcion", "description", "detalle", "concepto", "glosa", "memo",
        "movimiento", "detalle movimiento", "descripcion movimiento",
        "descripcion de transaccion", "transaction description", "narrative",
    ],
    "debito": [
        "debito", "debit", "debe", "cargo", "cargos", "retiro", "retiros",
        "egreso", "egresos", "withdrawal", "withdrawals", "valor -",
        "monto debito", "debito de transaccion", "cargo debito",
        "deb", "db", "dr",
    ],
    "credito": [
        "credito", "credit", "haber", "abono", "abonos", "deposito",
        "depositos", "ingreso", "ingresos", "deposit", "valor +",
        "monto credito", "credito de transaccion", "abono credito",
        "cr",
    ],
    "saldo": [
        "saldo", "balance", "saldo contable", "saldo disponible",
        "available balance", "running balance", "balance de transaccion",
        "saldo final", "book balance", "saldo real",
        "saldos",
    ],
    "monto": [
        "monto", "amount", "importe", "valor", "valor transaccion",
        "valor de transaccion", "monto transaccion", "transaction amount",
        "importe movimiento", "montos",
    ],
    "tipo": [
        "tipo", "naturaleza", "d/c", "dc", "dr/cr", "dr cr", "deb cred",
        "signo", "transaction type", "tipo movimiento",
        "deb./cred.", "deb./cred", "deb/cred",
    ],
}



# Soporte adicional bancos regionales:
# Cuscatlán utiliza Deb./Cred. + Valor + Saldo Real/Disponible.
# La clasificación final se realiza con el signo del movimiento (+/-).
GENERIC_SHEET_NAMES = {
    "hoja", "hoja1", "hoja 1", "sheet", "sheet1", "sheet 1",
    "movimientos", "transacciones", "estado de cuenta", "estado cuenta",
}

SUMMARY_PREFIXES = (
    "resumen", "total", "totales", "subtotal", "saldo inicial", "saldo anterior",
    "saldo final", "cantidad deb", "cantidad cred", "monto deb", "monto cred",
    "numero de transacciones", "no debitos", "no creditos", "summary",
    "beginning balance", "ending balance", "opening balance", "closing balance",
)

CREDIT_HINTS = (
    "credito", "credit", "abono", "deposito", "deposit", "ingreso", "haber",
    "transferencia recibida", "tef de", "ach credito", "cr", "c",
)

DEBIT_HINTS = (
    "debito", "debit", "cargo", "retiro", "egreso", "debe", "pago",
    "transferencia enviada", "ach debito", "dr", "d",
)

KNOWN_BANKS = [
    # Guatemala
    (["banco industrial", "bi"], "BANCO INDUSTRIAL"),
    (["g&t continental", "gyt continental", "g&t", "gyt"], "G&T"),
    (["bac credomatic", "banco de america central", "bac"], "BAC"),
    (["banrural", "banco de desarrollo rural", "ban"], "BANRURAL"),
    (["banco agricola mercantil", "bam"], "BAM"),
    (["banco promerica", "promerica"], "PROMERICA"),
    (["banco davivienda", "davivienda"], "DAVIVIENDA"),
    (["banco ficohsa", "ficohsa"], "FICOHSA"),
    # El Salvador / regionales
    (["banco agricola", "banco agrícola"], "BANCO AGRÍCOLA"),
    (["banco cuscatlan", "banco cuscatlán", "cuscatlan", "cuscatlán"], "CUSCATLÁN"),
    (["banco hipotecario"], "BANCO HIPOTECARIO"),
    (["banco azul", "abank"], "BANCO AZUL"),
    (["banco atlantida", "banco atlántida", "atlantida", "atlántida"], "BANCO ATLÁNTIDA"),
    (["banco promete", "banco integral"], "BANCO INTEGRAL"),
]


def clean_text(value):
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ").replace("\u200b", " ")

    # Elimina acentos y también tolera encabezados con codificación dañada
    # (ej.: "Descripci髇", "D閎ito", "Cr閐ito", "D bito").
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def compact_text(value):
    text = clean_text(value)
    return re.sub(r"[^a-z0-9+\-/& ]+", "", text).strip()


def parse_number(value):
    """Convierte números bancarios sin confundir 1,000 con 1.00.

    Devuelve None si la celda no contiene un número utilizable.
    """
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text in {"-", "--", "—", "–"}:
        return None

    original = clean_text(text)
    negative = False

    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    if text.endswith("-"):
        negative = True
        text = text[:-1]

    # Indicadores comunes de débito al lado del importe.
    if re.search(r"\b(db|dr|debit|debito)\b", original):
        negative = True

    text = text.replace("Q", "").replace("$", "").replace("€", "")
    text = re.sub(r"\b(GTQ|USD|EUR|SVC)\b", "", text, flags=re.I)
    text = text.replace("'", "").replace(" ", "")
    text = re.sub(r"[^0-9,\.\-+]", "", text)

    if not re.search(r"\d", text):
        return None

    sign = -1 if negative else 1
    if text.startswith("-"):
        sign = -1
    text = text.lstrip("+-")

    # Determina separador decimal según la posición del último separador.
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")

    elif "," in text:
        parts = text.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            text = parts[0] + "." + parts[1]
        elif len(parts) > 2 and len(parts[-1]) in (1, 2):
            text = "".join(parts[:-1]) + "." + parts[-1]
        else:
            text = "".join(parts)

    elif "." in text:
        parts = text.split(".")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            pass
        elif len(parts) > 2 and len(parts[-1]) in (1, 2):
            text = "".join(parts[:-1]) + "." + parts[-1]
        elif len(parts) == 2 and len(parts[1]) == 3:
            text = "".join(parts)

    try:
        return sign * abs(float(text))
    except Exception:
        return None


def valor_o_nd(value):
    """Devuelve 'N/D' cuando un dato realmente no existe.

    Se usa principalmente al escribir el Excel final. Los cálculos internos
    permanecen numéricos para no romper sumas, reportes ni el tablero.
    """
    if value is None:
        return "N/D"
    if isinstance(value, str) and not value.strip():
        return "N/D"
    return value


def clean_number(value):
    parsed = parse_number(value)
    return parsed if parsed is not None else 0.0



def detectar_moneda_textual(
    rows=None,
    worksheet_title="",
    nombre_archivo=""
):
    """
    Detecta moneda usando SOLO señales textuales reales del archivo:
    códigos (USD/GTQ/EUR...), nombres de moneda y símbolos escritos
    junto a importes.

    No usa number_format de Excel, porque un archivo transformado puede
    conservar un formato viejo (por ejemplo Q) aunque el estado indique USD.
    """

    scores = {
        "GTQ": 0,
        "USD": 0,
        "EUR": 0,
        "SVC": 0,
        "HNL": 0,
        "CRC": 0,
        "NIO": 0,
        "MXN": 0,
        "GBP": 0,
    }

    def analizar(value, peso=1):
        if value in (None, ""):
            return

        raw = str(value).strip()
        if not raw:
            return

        normal = clean_text(raw)

        patrones = {
            "GTQ": (
                r"\bGTQ\b|\bQTZ\b|\bQUETZAL(?:ES)?\b",
                r"(?<![A-Za-z])Q\s*[-+]?\s*\d",
            ),
            "USD": (
                r"\bUSD\b|\bD[OÓ]LAR(?:ES)?\b|\bDOLLAR(?:S)?\b",
                r"US\s*\$\s*[-+]?\s*\d",
            ),
            "EUR": (
                r"\bEUR\b|\bEURO(?:S)?\b",
                r"€\s*[-+]?\s*\d",
            ),
            "SVC": (
                r"\bSVC\b|\bCOL[ÓO]N(?:ES)?\s+SALVADORE[ÑN]O(?:S)?\b",
                None,
            ),
            "HNL": (
                r"\bHNL\b|\bLEMPIRA(?:S)?\b|\bHONDURAS\b",
                None,
            ),
            "CRC": (
                r"\bCRC\b|\bCOL[ÓO]N(?:ES)?\s+COSTARRICENSE(?:S)?\b",
                None,
            ),
            "NIO": (
                r"\bNIO\b|\bC[ÓO]RDOBA(?:S)?\b",
                r"C\$\s*[-+]?\s*\d",
            ),
            "MXN": (
                r"\bMXN\b|\bPESO(?:S)?\s+MEXICANO(?:S)?\b",
                r"MX\$\s*[-+]?\s*\d",
            ),
            "GBP": (
                r"\bGBP\b|\bLIBRA(?:S)?\s+ESTERLINA(?:S)?\b|\bPOUND(?:S)?\b",
                r"£\s*[-+]?\s*\d",
            ),
        }

        for codigo, (patron_texto, patron_simbolo) in patrones.items():
            if re.search(patron_texto, raw, re.I):
                scores[codigo] += 200 * peso

            if patron_simbolo and re.search(
                patron_simbolo,
                raw,
                re.I
            ):
                scores[codigo] += 120 * peso

        # $ sin prefijo se toma como USD solo cuando está escrito en el
        # contenido de una celda, no cuando viene de un number_format.
        if re.search(r"(?<![A-Za-z])\$\s*[-+]?\s*\d", raw):
            scores["USD"] += 80 * peso

    analizar(nombre_archivo, 3)
    analizar(worksheet_title, 3)

    rows = rows or []

    for row in rows[:60]:
        for value in row:
            analizar(value, 1)

    ranking = sorted(
        scores.items(),
        key=lambda item: item[1],
        reverse=True
    )

    if not ranking or ranking[0][1] <= 0:
        return "N/D"

    if (
        len(ranking) > 1
        and ranking[0][1] == ranking[1][1]
    ):
        return "N/D"

    return ranking[0][0]


def detectar_moneda_contexto_archivo(
    workbook,
    nombre_archivo=""
):
    """
    Obtiene una moneda de respaldo para el libro completo.

    Prioridad:
    1. Moneda escrita explícitamente en los estados.
    2. País/empresa escrito en el archivo.
    3. Bancos fuertemente asociados al país como último respaldo.

    Esta moneda SOLO se usa en hojas donde no haya una moneda textual
    explícita. Por eso una cuenta USD en Guatemala puede seguir siendo USD
    si el estado de cuenta lo indica expresamente.
    """

    monedas_textuales = set()
    contexto = [str(nombre_archivo or "")]

    for worksheet in workbook.worksheets:

        rows_contexto = []

        try:
            for row_index, row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1
            ):
                if row_index > 40:
                    break

                valores = list(row)
                rows_contexto.append(valores)

                for value in valores:
                    if value not in (None, ""):
                        contexto.append(str(value))

        except Exception:
            pass

        contexto.append(str(worksheet.title))

        moneda_textual = detectar_moneda_textual(
            rows_contexto,
            worksheet.title,
            nombre_archivo
        )

        if moneda_textual not in {"N/D", "MULTI"}:
            monedas_textuales.add(moneda_textual)

    # Si todo lo escrito explícitamente coincide, esa es la mejor
    # moneda de respaldo del archivo.
    if len(monedas_textuales) == 1:
        return next(iter(monedas_textuales))

    # Si hay dos monedas explícitas reales, no forzar una sola.
    if len(monedas_textuales) > 1:
        return "MULTI"

    texto_contexto = clean_text(
        " | ".join(contexto)
    )

    # País / empresa como respaldo.
    if "el salvador" in texto_contexto:
        return "USD"

    if "guatemala" in texto_contexto:
        return "GTQ"

    if "honduras" in texto_contexto:
        return "HNL"

    if "costa rica" in texto_contexto:
        return "CRC"

    if "nicaragua" in texto_contexto:
        return "NIO"

    if "mexico" in texto_contexto:
        return "MXN"

    # Último respaldo por bancos muy representativos del país.
    senales_gt = (
        "banco industrial",
        "banrural",
        "g&t",
        "gyt",
        "banco agricola mercantil",
        " bam ",
    )

    senales_sv = (
        "banco agricola",
        "banco agricola",
        "cuscatlan",
        "cuscatlan",
    )

    gt = sum(
        1
        for señal in senales_gt
        if señal in texto_contexto
    )

    sv = sum(
        1
        for señal in senales_sv
        if señal in texto_contexto
    )

    if gt > 0 and gt > sv:
        return "GTQ"

    if sv > 0 and sv > gt:
        return "USD"

    return "N/D"


def detectar_moneda(worksheet, rows=None, header_index=None):
    """
    Detecta la moneda únicamente para PRESENTACIÓN en el dashboard.
    NO convierte importes, NO separa totales y NO modifica cálculos.
    """
    scores = {
        "GTQ": 0,   # Quetzal
        "USD": 0,   # Dólar estadounidense
        "EUR": 0,   # Euro
        "SVC": 0,   # Colón salvadoreño
        "HNL": 0,   # Lempira
        "CRC": 0,   # Colón costarricense
        "NIO": 0,   # Córdoba
        "MXN": 0,   # Peso mexicano
        "GBP": 0,   # Libra esterlina
    }

    def analizar_texto(value, peso=1):
        if value in (None, ""):
            return

        raw = str(value).strip()
        if not raw:
            return

        normal = clean_text(raw)

        # Códigos y nombres explícitos
        if re.search(r"\bGTQ\b|\bQTZ\b|\bQUETZAL(?:ES)?\b", normal, re.I):
            scores["GTQ"] += 100 * peso

        if re.search(r"\bUSD\b|\bD[OÓ]LAR(?:ES)?\b|\bDOLLAR(?:S)?\b", raw, re.I):
            scores["USD"] += 100 * peso

        if re.search(r"\bEUR\b|\bEURO(?:S)?\b", normal, re.I):
            scores["EUR"] += 100 * peso

        if re.search(r"\bSVC\b", normal, re.I):
            scores["SVC"] += 100 * peso

        if re.search(r"\bHNL\b|\bLEMPIRA(?:S)?\b", normal, re.I):
            scores["HNL"] += 100 * peso

        if re.search(r"\bCRC\b|\bCOLON(?:ES)? COSTARRICENSE(?:S)?\b", normal, re.I):
            scores["CRC"] += 100 * peso

        if re.search(r"\bNIO\b|\bCORDOBA(?:S)?\b", normal, re.I):
            scores["NIO"] += 100 * peso

        if re.search(r"\bMXN\b|\bPESO(?:S)? MEXICANO(?:S)?\b", normal, re.I):
            scores["MXN"] += 100 * peso

        if re.search(r"\bGBP\b|\bLIBRA(?:S)? ESTERLINA(?:S)?\b|\bPOUND(?:S)?\b", normal, re.I):
            scores["GBP"] += 100 * peso

        # Símbolos junto a importes
        if re.search(r"(?<![A-Za-z])Q\s*[-+]?\s*\d", raw, re.I):
            scores["GTQ"] += 50 * peso

        if re.search(r"US\s*\$\s*[-+]?\s*\d", raw, re.I):
            scores["USD"] += 60 * peso
        elif re.search(r"\$\s*[-+]?\s*\d", raw):
            scores["USD"] += 25 * peso

        if re.search(r"€\s*[-+]?\s*\d", raw):
            scores["EUR"] += 60 * peso

        if re.search(r"£\s*[-+]?\s*\d", raw):
            scores["GBP"] += 60 * peso

        if re.search(r"₡\s*[-+]?\s*\d", raw):
            # ₡ puede corresponder a CRC o SVC. Si el texto no aclara,
            # se favorece CRC por uso actual del símbolo.
            scores["CRC"] += 45 * peso

        if re.search(r"C\$\s*[-+]?\s*\d", raw, re.I):
            scores["NIO"] += 45 * peso

        if re.search(r"MX\$\s*[-+]?\s*\d", raw, re.I):
            scores["MXN"] += 55 * peso

    analizar_texto(worksheet.title, 2)

    rows = rows or []
    limite = min(len(rows), max((header_index or 0) + 10, 40))

    for row in rows[:limite]:
        for value in row:
            analizar_texto(value, 1)

    # También revisar formatos de celdas de Excel.
    #
    # IMPORTANTE:
    # En algunos archivos bancarios abiertos con read_only=True,
    # openpyxl puede devolver None en worksheet.max_row/max_column.
    # Por eso NO usamos min(worksheet.max_row, ...).
    #
    # Recorremos de forma segura un máximo de 80 filas y 40 columnas.
    try:
        for row_index, row in enumerate(
            worksheet.iter_rows(),
            start=1
        ):
            if row_index > 80:
                break

            for column_index, cell in enumerate(
                row,
                start=1
            ):
                if column_index > 40:
                    break

                fmt = str(
                    getattr(
                        cell,
                        "number_format",
                        ""
                    ) or ""
                )

                fmt_up = fmt.upper()

                if (
                    "GTQ" in fmt_up
                    or "QTZ" in fmt_up
                    or re.search(
                        r'(^|[^A-Z])Q([^A-Z]|$)',
                        fmt_up
                    )
                ):
                    scores["GTQ"] += 10

                if (
                    "USD" in fmt_up
                    or "US$" in fmt_up
                ):
                    scores["USD"] += 12

                elif "$" in fmt:
                    scores["USD"] += 4

                if (
                    "EUR" in fmt_up
                    or "€" in fmt
                ):
                    scores["EUR"] += 12

                if "SVC" in fmt_up:
                    scores["SVC"] += 12

                if "HNL" in fmt_up:
                    scores["HNL"] += 12

                if "CRC" in fmt_up:
                    scores["CRC"] += 12

                if "NIO" in fmt_up or "C$" in fmt_up:
                    scores["NIO"] += 12

                if "MXN" in fmt_up or "MX$" in fmt_up:
                    scores["MXN"] += 12

                if "GBP" in fmt_up or "£" in fmt:
                    scores["GBP"] += 12

                if "₡" in fmt and "SVC" not in fmt_up:
                    scores["CRC"] += 8

    except Exception as currency_format_error:
        # La moneda es un dato de presentación.
        # Nunca debe impedir que el estado de cuenta sea procesado.
        print(
            "Advertencia detectando formato de moneda en hoja",
            worksheet.title,
            ":",
            str(currency_format_error)
        )

    ranking = sorted(scores.items(), key=lambda x: x[1], reverse=True)

    if not ranking or ranking[0][1] <= 0:
        return "N/D"

    if len(ranking) > 1 and ranking[0][1] == ranking[1][1]:
        return "N/D"

    return ranking[0][0]



def normalizar_codigo_moneda(moneda):
    """
    Normaliza códigos de moneda únicamente para formato visual.
    No convierte valores ni modifica cálculos.
    """
    texto = str(moneda or "").strip().upper()

    aliases = {
        "GTQ": "GTQ",
        "QTZ": "GTQ",
        "Q": "GTQ",
        "QUETZAL": "GTQ",
        "QUETZALES": "GTQ",
        "USD": "USD",
        "US$": "USD",
        "$": "USD",
        "DOLAR": "USD",
        "DÓLAR": "USD",
        "DOLARES": "USD",
        "DÓLARES": "USD",
        "EUR": "EUR",
        "EURO": "EUR",
        "EUROS": "EUR",
        "SVC": "SVC",
        "COLON SALVADORENO": "SVC",
        "COLÓN SALVADOREÑO": "SVC",
        "HNL": "HNL",
        "LEMPIRA": "HNL",
        "LEMPIRAS": "HNL",
        "CRC": "CRC",
        "COLON COSTARRICENSE": "CRC",
        "COLÓN COSTARRICENSE": "CRC",
        "NIO": "NIO",
        "CORDOBA": "NIO",
        "CÓRDOBA": "NIO",
        "CORDOBAS": "NIO",
        "CÓRDOBAS": "NIO",
        "MXN": "MXN",
        "PESO MEXICANO": "MXN",
        "PESOS MEXICANOS": "MXN",
        "GBP": "GBP",
        "LIBRA ESTERLINA": "GBP",
        "LIBRAS ESTERLINAS": "GBP",
    }

    codigos_validos = {
        "GTQ", "USD", "EUR", "SVC",
        "HNL", "CRC", "NIO", "MXN", "GBP"
    }

    return aliases.get(
        texto,
        texto if texto in codigos_validos else "N/D"
    )


def formato_moneda_excel(moneda):
    """
    Devuelve el formato numérico de Excel según la moneda detectada.
    Se cambia únicamente la presentación, nunca el valor.
    """
    codigo = normalizar_codigo_moneda(moneda)

    formatos = {
        "GTQ": '"Q" #,##0.00;[Red]-"Q" #,##0.00',
        "USD": '"$" #,##0.00;[Red]-"$" #,##0.00',
        "EUR": '"€" #,##0.00;[Red]-"€" #,##0.00',
        "SVC": '"₡" #,##0.00;[Red]-"₡" #,##0.00',
        "HNL": '"L" #,##0.00;[Red]-"L" #,##0.00',
        "CRC": '"₡" #,##0.00;[Red]-"₡" #,##0.00',
        "NIO": '"C$" #,##0.00;[Red]-"C$" #,##0.00',
        "MXN": '"MX$" #,##0.00;[Red]-"MX$" #,##0.00',
        "GBP": '"£" #,##0.00;[Red]-"£" #,##0.00',
    }

    return formatos.get(
        codigo,
        '#,##0.00;[Red]-#,##0.00'
    )


def moneda_unica_transacciones(transactions):
    """
    Devuelve una moneda solo cuando todas las transacciones identificadas
    corresponden a la misma moneda.

    Si hubiera varias monedas, devuelve MULTI para no colocar un símbolo
    incorrecto en totales consolidados.
    """
    monedas = {
        normalizar_codigo_moneda(
            transaction.get("moneda")
        )
        for transaction in transactions
        if normalizar_codigo_moneda(
            transaction.get("moneda")
        ) != "N/D"
    }

    if len(monedas) == 1:
        return next(iter(monedas))

    if len(monedas) > 1:
        return "MULTI"

    return "N/D"


def aplicar_formato_moneda_excel(
    workbook,
    transactions
):
    """
    Aplica el símbolo de moneda al Excel FINAL.

    IMPORTANTE:
    - NO cambia importes.
    - NO convierte monedas.
    - NO cambia fórmulas.
    - NO cambia sumas.
    - Solo modifica number_format de celdas monetarias.
    """

    if not transactions:
        return

    moneda_general = moneda_unica_transacciones(
        transactions
    )

    formato_general = formato_moneda_excel(
        moneda_general
    )

    # --------------------------------------------------------
    # MAPAS DE MONEDA POR CUENTA
    # --------------------------------------------------------

    moneda_por_cuenta_original = {}
    moneda_por_cuenta_corta = {}

    for transaction in transactions:

        moneda = normalizar_codigo_moneda(
            transaction.get("moneda")
        )

        banco_original = str(
            transaction.get("banco", "")
        ).strip()

        banco_resumido = banco_corto(
            banco_original
        )

        cuenta = str(
            transaction.get("cuenta", "")
        ).strip()

        clave_original = (
            banco_original,
            cuenta
        )

        clave_corta = (
            banco_resumido,
            cuenta
        )

        if moneda != "N/D":
            moneda_por_cuenta_original[
                clave_original
            ] = moneda

            moneda_por_cuenta_corta[
                clave_corta
            ] = moneda

    # --------------------------------------------------------
    # 1) ESTADOS CONSOLIDADOS
    #    Débito, Crédito, Saldo
    # --------------------------------------------------------

    if SHEET_ESTADOS in workbook.sheetnames:

        ws = workbook[
            SHEET_ESTADOS
        ]

        for row_number, transaction in enumerate(
            transactions,
            start=2
        ):

            formato = formato_moneda_excel(
                transaction.get(
                    "moneda"
                )
            )

            for column in (
                6,
                7,
                8
            ):

                cell = obtener_celda_segura(
                    ws,
                    row_number,
                    column
                )

                if cell:
                    cell.number_format = formato

    # --------------------------------------------------------
    # 2) SALDOS POR CUENTA
    #    Saldo inicial y saldo final
    # --------------------------------------------------------

    if SHEET_SALDOS in workbook.sheetnames:

        ws = workbook[
            SHEET_SALDOS
        ]

        cuentas = {}

        for transaction in transactions:

            clave = (
                transaction["banco"],
                transaction["cuenta"]
            )

            cuentas.setdefault(
                clave,
                []
            ).append(
                transaction
            )

        row_number = 5

        for (
            banco,
            cuenta
        ), movimientos in cuentas.items():

            moneda = moneda_por_cuenta_original.get(
                (
                    str(banco).strip(),
                    str(cuenta).strip()
                ),
                "N/D"
            )

            formato = formato_moneda_excel(
                moneda
            )

            for column in (
                3,
                4
            ):

                cell = obtener_celda_segura(
                    ws,
                    row_number,
                    column
                )

                if cell:
                    cell.number_format = formato

            row_number += 1

    # --------------------------------------------------------
    # 3) REPORTE CRÉDITOS DIARIOS
    #    Cada cuenta conserva la moneda de su banco.
    # --------------------------------------------------------

    if SHEET_REPORTE in workbook.sheetnames:

        ws = workbook[
            SHEET_REPORTE
        ]

        cuentas = sorted(
            {
                (
                    banco_corto(
                        transaction["banco"]
                    ),
                    str(
                        transaction["cuenta"]
                    )
                )
                for transaction in transactions
            },
            key=lambda item: (
                item[0],
                item[1]
            )
        )

        # Fechas realmente procesadas.
        fechas = {
            (
                transaction["fecha"].date()
                if isinstance(
                    transaction.get("fecha"),
                    datetime
                )
                else transaction.get("fecha")
            )
            for transaction in transactions
            if isinstance(
                transaction.get("fecha"),
                (datetime, date)
            )
        }

        last_data_row = (
            4 + len(fechas)
            if fechas
            else 4
        )

        # Las cuentas reales siempre empiezan desde la columna B
        # en el mismo orden usado al crear el reporte.
        for offset, (
            banco,
            cuenta
        ) in enumerate(
            cuentas
        ):

            column = 2 + offset

            moneda = moneda_por_cuenta_corta.get(
                (
                    banco,
                    cuenta
                ),
                "N/D"
            )

            formato = formato_moneda_excel(
                moneda
            )

            for row in range(
                5,
                last_data_row + 1
            ):

                cell = obtener_celda_segura(
                    ws,
                    row,
                    column
                )

                if cell:
                    cell.number_format = formato

        # Detectar columnas de TOTAL y CUENTAS CON ABONO.
        total_col = None
        count_col = None

        for col in range(
            2,
            ws.max_column + 1
        ):

            header = clean_text(
                ws.cell(
                    row=4,
                    column=col
                ).value
            )

            if (
                total_col is None
                and "total" in header
                and (
                    "credito" in header
                    or "credit" in header
                )
            ):
                total_col = col

            if (
                count_col is None
                and "cuentas" in header
                and "abono" in header
            ):
                count_col = col

        # El TOTAL recibe símbolo solo si existe una moneda única.
        if total_col is not None:

            for row in range(
                5,
                last_data_row + 1
            ):

                cell = obtener_celda_segura(
                    ws,
                    row,
                    total_col
                )

                if cell:
                    cell.number_format = formato_general

        # La cantidad de cuentas nunca lleva moneda.
        if count_col is not None:

            for row in range(
                5,
                last_data_row + 1
            ):

                cell = obtener_celda_segura(
                    ws,
                    row,
                    count_col
                )

                if cell:
                    cell.number_format = "0"

    # --------------------------------------------------------
    # 4) TABLERO CRÉDITOS
    # --------------------------------------------------------

    if SHEET_TABLERO in workbook.sheetnames:

        ws = workbook[
            SHEET_TABLERO
        ]

        # --------------------------------------------------------
        # KPIs SUPERIORES DEL TABLERO
        # --------------------------------------------------------
        # La plantilla trae estos indicadores con formato monetario fijo.
        # Buscamos los títulos dinámicamente y aplicamos la moneda detectada
        # únicamente a la celda de valor que está inmediatamente debajo.
        #
        # CRÉDITOS EN RANGO  -> moneda
        # PROMEDIO DIARIO    -> moneda
        # MAYOR DÍA          -> moneda
        # CUENTAS CON ABONO  -> entero, sin símbolo
        # --------------------------------------------------------

        titulos_monetarios = (
            "creditos en rango",
            "promedio diario",
            "mayor dia",
            "saldo disponible",
        )

        titulo_cantidad = "cuentas con abono"

        for row in range(
            1,
            min(ws.max_row, 14) + 1
        ):

            for column in range(
                1,
                ws.max_column + 1
            ):

                cell = ws.cell(
                    row=row,
                    column=column
                )

                if es_celda_combinada(cell):
                    continue

                texto = clean_text(
                    cell.value
                )

                if not texto:
                    continue

                # Indicadores monetarios.
                if any(
                    titulo in texto
                    for titulo in titulos_monetarios
                ):

                    value_cell = obtener_celda_segura(
                        ws,
                        row + 1,
                        column
                    )

                    if value_cell:
                        value_cell.number_format = formato_general

                # Indicador de cantidad.
                elif titulo_cantidad in texto:

                    value_cell = obtener_celda_segura(
                        ws,
                        row + 1,
                        column
                    )

                    if value_cell:
                        value_cell.number_format = "0"

        # Evolución: Crédito y Variación.
        # Si todo el archivo corresponde a una moneda, se usa su símbolo.
        for row in range(
            15,
            ws.max_row + 1
        ):

            for column in (
                2,
                3
            ):

                cell = obtener_celda_segura(
                    ws,
                    row,
                    column
                )

                if cell:
                    cell.number_format = formato_general

        # Créditos por cuenta: columna H.
        cuentas = sorted(
            {
                (
                    banco_corto(
                        transaction["banco"]
                    ),
                    str(
                        transaction["cuenta"]
                    )
                )
                for transaction in transactions
            },
            key=lambda item: (
                item[0],
                item[1]
            )
        )

        for index, (
            banco,
            cuenta
        ) in enumerate(
            cuentas,
            start=15
        ):

            moneda = moneda_por_cuenta_corta.get(
                (
                    banco,
                    cuenta
                ),
                "N/D"
            )

            cell = obtener_celda_segura(
                ws,
                index,
                8
            )

            if cell:
                cell.number_format = formato_moneda_excel(
                    moneda
                )

        # Saldo Disponible y Total usan la moneda única del libro.
        saldo_disponible_cell = obtener_celda_segura(ws, 8, 9)
        if saldo_disponible_cell:
            saldo_disponible_cell.number_format = formato_general

        total_row = 15 + len(cuentas)
        total_cell = obtener_celda_segura(ws, total_row, 8)
        if total_cell:
            total_cell.number_format = formato_general

        # Los ejes monetarios de las gráficas también se adaptan
        # cuando el libro utiliza una única moneda.
        charts = getattr(
            ws,
            "_charts",
            []
        )

        if moneda_general not in (
            "N/D",
            "MULTI"
        ):

            chart_format = formato_moneda_excel(
                moneda_general
            )

            for chart in charts:

                try:
                    if hasattr(
                        chart,
                        "y_axis"
                    ):
                        chart.y_axis.numFmt = chart_format
                except Exception:
                    pass


def _parse_text_date(text, expected_month=None, expected_year=None):
    text = str(text).strip()
    if not text:
        return None

    # Limpia hora cuando viene junto a la fecha.
    text = re.sub(r"\s+\d{1,2}:\d{2}(:\d{2})?.*$", "", text)

    # ISO primero (no ambiguo).
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass

    # dd/mm es el estándar principal para Guatemala y El Salvador.
    ddmm = None
    mmdd = None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            ddmm = datetime.strptime(text, fmt)
            break
        except Exception:
            pass

    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%m.%d.%Y", "%m/%d/%y", "%m-%d-%y"):
        try:
            mmdd = datetime.strptime(text, fmt)
            break
        except Exception:
            pass

    if expected_month:
        candidates = [d for d in (ddmm, mmdd) if d is not None]
        for d in candidates:
            if d.month == expected_month and (not expected_year or d.year == expected_year):
                return d

    return ddmm or mmdd


def clean_date(value, expected_month=None, expected_year=None):
    if value is None or value == "":
        return ""

    parsed = None

    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day)
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        # Los seriales de Excel razonables están aprox. entre 20,000 y 80,000.
        if 20000 <= float(value) <= 80000:
            try:
                from openpyxl.utils.datetime import from_excel
                parsed = from_excel(value)
            except Exception:
                parsed = None
    else:
        parsed = _parse_text_date(value, expected_month, expected_year)

    if not isinstance(parsed, datetime):
        return ""

    # Algunos exports bancarios latinoamericanos guardan 04/08 como serial
    # equivalente a 08/04. Si conocemos el mes del estado de cuenta, se corrige
    # de forma genérica (sin depender del banco).
    if expected_month and parsed.month != expected_month and parsed.day == expected_month:
        try:
            year = expected_year or parsed.year
            parsed = datetime(year, expected_month, parsed.month)
        except Exception:
            pass

    if expected_year and parsed.year != expected_year:
        # Solo ajustar años abreviados/erróneos cuando el mes coincide y la
        # diferencia es claramente de interpretación del periodo.
        if parsed.month == expected_month and abs(parsed.year - expected_year) <= 1:
            try:
                parsed = parsed.replace(year=expected_year)
            except Exception:
                pass

    return parsed


def _similarity(a, b):
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if b in a or a in b:
        short = min(len(a), len(b))
        long = max(len(a), len(b))
        return max(0.86, short / max(long, 1))
    return SequenceMatcher(None, a, b).ratio()


def _header_field(value):
    text = compact_text(value)
    if not text:
        return None, 0.0

    best_field = None
    best_score = 0.0

    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_clean = compact_text(alias)
            score = _similarity(text, alias_clean)

            # Coincidencias por raíces robustas a caracteres dañados.
            if field == "descripcion" and ("descrip" in text or "concept" in text or "detalle" in text):
                score = max(score, 0.98)
            elif field == "debito" and (
                any(root in text for root in ("deb", "debe", "cargo", "reti", "egres", "dbit"))
                or text.startswith("dito")
            ):
                score = max(score, 0.96)
            elif field == "credito" and (
                any(root in text for root in ("cred", "haber", "abono", "deposit", "ingres", "crdit"))
                or text.startswith("crito")
            ):
                score = max(score, 0.96)
            elif field == "fecha" and (text == "fecha" or text.startswith("fecha ") or " date" in f" {text}"):
                score = max(score, 0.98)
            elif field == "saldo" and ("saldo" in text or "balance" in text):
                score = max(score, 0.98)
            elif field == "referencia" and any(root in text for root in ("refer", "no doc", "document", "secuencial", "trace")):
                score = max(score, 0.97)
            elif field == "codigo" and any(root in text for root in ("codigo", "cigo", "transaction code")):
                score = max(score, 0.95)
            elif field == "monto" and any(root in text for root in ("monto", "importe", "amount")):
                score = max(score, 0.95)

            if score > best_score:
                best_field = field
                best_score = score

    if best_score >= 0.78:
        return best_field, best_score
    return None, best_score


def detectar_periodo(rows, sheet_name=""):
    """Devuelve (mes, año) del estado de cuenta cuando puede inferirse."""
    month = None
    year = None

    title = clean_text(sheet_name)

    # Ej.: "BAC 08.26", "Cuenta_08-2026", "2026-08".
    m = re.search(r"(?<!\d)(0?[1-9]|1[0-2])[.\-_/ ](20\d{2}|\d{2})(?!\d)", title)
    if m:
        month = int(m.group(1))
        y = int(m.group(2))
        year = 2000 + y if y < 100 else y

    if month is None:
        m = re.search(r"(?<!\d)(20\d{2})[.\-_/ ](0?[1-9]|1[0-2])(?!\d)", title)
        if m:
            year = int(m.group(1))
            month = int(m.group(2))

    # Busca periodos explícitos en las primeras filas.
    for row in rows[:25]:
        joined = " | ".join(str(v) for v in row if v not in (None, ""))
        normalized = clean_text(joined)
        if any(k in normalized for k in ("fecha inicial", "fecha final", "periodo", "desde", "hasta", "corte")):
            dates = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", joined)
            for dtext in dates:
                d = _parse_text_date(dtext)
                if d:
                    month = month or d.month
                    year = year or d.year
                    if month == d.month:
                        year = d.year
                        break

    return month, year



def detectar_banco_por_nombre_archivo(nombre_archivo):
    """
    Detecta el banco usando primero el nombre REAL del archivo recibido.

    Se usan alias conservadores para los nombres operativos de los estados
    salvadoreños. Esto evita OCR innecesario en Make/Render y reduce falsos
    positivos. Si no hay una coincidencia clara, devuelve None y el parser
    continúa con hoja/celdas y, como último respaldo, OCR.
    """
    texto = clean_text(os.path.basename(str(nombre_archivo or "")))

    if not texto:
        return None

    # Primero intentar el catálogo general con nombres completos.
    banco_catalogo = _buscar_banco_catalogo_en_texto(texto)
    if banco_catalogo:
        return banco_catalogo

    # Alias de archivo usados en los estados de cuenta de El Salvador.
    # Se exigen límites de palabra para evitar coincidencias accidentales.
    patrones_archivo = (
        (r"\bpromerica\b", "PROMERICA"),
        (r"\bdav(?:ivienda)?\b", "DAVIVIENDA"),
        (r"\bcusca(?:tlan)?\b", "CUSCATLÁN"),
        (r"\bbac\b", "BAC"),
        (r"\bagricola\b", "BANCO AGRÍCOLA"),
    )

    for patron, banco in patrones_archivo:
        if re.search(patron, texto, re.I):
            return banco

    return None

def detectar_banco(sheet_name, rows):
    """
    Detecta bancos únicamente contra KNOWN_BANKS.

    REGLA IMPORTANTE:
    - Nunca inventa un banco a partir de la palabra "Banco".
    - Nunca usa Guatemala / El Salvador / GTQ / USD como nombre de banco.
    - Nunca usa el nombre de la hoja como banco si no coincide con el catálogo.
    - Si no existe coincidencia textual, devuelve BANCO NO IDENTIFICADO para que
      posteriormente pueda actuar el respaldo OCR.
    """

    title = clean_text(sheet_name)

    # 1) Buscar banco conocido en el nombre de la hoja.
    banco_catalogo = _buscar_banco_catalogo_en_texto(title)
    if banco_catalogo:
        return banco_catalogo

    # 2) Buscar únicamente bancos conocidos dentro de las primeras filas.
    # No se acepta texto genérico del tipo "Banco Guatemala", "Banco", etc.
    for row in rows[:40]:
        for value in row:
            if value in (None, ""):
                continue

            banco_catalogo = _buscar_banco_catalogo_en_texto(value)
            if banco_catalogo:
                return banco_catalogo

    # 3) Sin coincidencia real: NO fabricar un nombre.
    return "BANCO NO IDENTIFICADO"


# ============================================================
# OCR DE RESPALDO PARA LOGOS / IMÁGENES DE BANCOS EN EXCEL
# ============================================================
#
# Esta lógica NO reemplaza la detección actual por texto. Solamente se activa
# cuando el nombre de la hoja y las celdas no permiten reconocer un banco del
# catálogo KNOWN_BANKS. Así los archivos que ya funcionan siguen usando el
# mismo camino rápido y el OCR se reserva para estados que traen el banco como
# imagen (por ejemplo un logo BAC CREDOMATIC).
#
# Dependencias opcionales en requirements.txt:
#   Pillow
#   numpy
#   rapidocr-onnxruntime
#
# Si estas dependencias no están disponibles, el parser NO se cae: conserva el
# comportamiento anterior y deja una advertencia en el log de Render.

_OCR_ENGINE = None
_OCR_ENGINE_ERROR = None

# OCR seguro para Render/Make: solo se usa como respaldo cuando el banco no
# aparece en texto. Se limita tanto por cantidad de imágenes como por tiempo
# total para evitar que /process-excel se prolongue innecesariamente.
OCR_MAX_IMAGES_PER_SHEET = 3
OCR_MAX_SECONDS_PER_FILE = 12


def _buscar_banco_catalogo_en_texto(texto):
    """Busca un banco de KNOWN_BANKS dentro de cualquier texto libre."""
    normal = clean_text(texto)

    if not normal:
        return None

    for aliases, canonical in KNOWN_BANKS:
        for alias in aliases:
            a = clean_text(alias)

            if not a:
                continue

            # Alias cortos como BAC, BI o BAM deben respetar límites de palabra
            # para no coincidir accidentalmente dentro de otra palabra.
            if len(a) <= 4 and re.fullmatch(r"[a-z0-9]+", a):
                if re.search(rf"\b{re.escape(a)}\b", normal):
                    return canonical
            elif a in normal:
                return canonical

    return None


def _obtener_motor_ocr():
    """Inicializa RapidOCR una sola vez por proceso de Render."""
    global _OCR_ENGINE, _OCR_ENGINE_ERROR

    if _OCR_ENGINE is not None:
        return _OCR_ENGINE

    if _OCR_ENGINE_ERROR is not None:
        return None

    try:
        from rapidocr_onnxruntime import RapidOCR

        _OCR_ENGINE = RapidOCR()
        return _OCR_ENGINE

    except Exception as error:
        _OCR_ENGINE_ERROR = str(error)
        print(
            "OCR bancario no disponible. "
            "Agrega Pillow, numpy y rapidocr-onnxruntime a requirements.txt:",
            str(error)
        )
        return None


def _extraer_textos_resultado_ocr(resultado):
    """Tolera distintas versiones de RapidOCR y devuelve una lista de textos."""
    if resultado is None:
        return []

    # RapidOCR reciente puede devolver un objeto con .txts
    if hasattr(resultado, "txts"):
        try:
            return [
                str(texto)
                for texto in (resultado.txts or [])
                if str(texto or "").strip()
            ]
        except Exception:
            pass

    # Versiones clásicas suelen devolver (result, elapsed)
    if isinstance(resultado, tuple):
        resultado = resultado[0] if resultado else None

    if resultado is None:
        return []

    textos = []

    try:
        for item in resultado:
            if isinstance(item, (list, tuple)) and len(item) >= 2:
                texto = str(item[1] or "").strip()
                if texto:
                    textos.append(texto)
            elif isinstance(item, str) and item.strip():
                textos.append(item.strip())
    except Exception:
        pass

    return textos


def _ocr_texto_imagen_bytes(image_bytes):
    """Lee texto visual dentro de una imagen y devuelve una cadena unificada."""
    engine = _obtener_motor_ocr()

    if engine is None or not image_bytes:
        return ""

    try:
        from PIL import Image, ImageOps
        import numpy as np

        imagen = Image.open(
            io.BytesIO(image_bytes)
        ).convert("RGB")

        # Logos pequeños se reconocen mejor al ampliarlos. Se limita el tamaño
        # para no disparar el consumo de memoria en Render.
        width, height = imagen.size
        lado_mayor = max(width, height)

        if lado_mayor and lado_mayor < 1200:
            escala = min(
                3.0,
                max(1.0, 1200.0 / float(lado_mayor))
            )

            if escala > 1.05:
                nuevo_tamano = (
                    max(1, int(width * escala)),
                    max(1, int(height * escala))
                )

                try:
                    resample = Image.Resampling.LANCZOS
                except AttributeError:
                    resample = Image.LANCZOS

                imagen = imagen.resize(
                    nuevo_tamano,
                    resample
                )

        variantes = [imagen]

        try:
            variantes.append(
                ImageOps.autocontrast(imagen)
            )
        except Exception:
            pass

        textos_totales = []

        for variante in variantes:
            try:
                salida = engine(
                    np.asarray(variante)
                )
                textos = _extraer_textos_resultado_ocr(
                    salida
                )
                textos_totales.extend(textos)

                # Si ya apareció un banco conocido, no gastar otra pasada.
                texto_parcial = " | ".join(textos_totales)
                if _buscar_banco_catalogo_en_texto(texto_parcial):
                    break

            except Exception as error:
                print(
                    "Advertencia OCR en una imagen:",
                    str(error)
                )

        return " | ".join(textos_totales)

    except Exception as error:
        print(
            "No se pudo preparar imagen para OCR:",
            str(error)
        )
        return ""


def _hojas_que_requieren_ocr(workbook):
    """Detecta qué hojas no tienen un banco conocido en título/celdas."""
    pendientes = []

    for worksheet in workbook.worksheets:
        rows_contexto = []

        try:
            for row_index, row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1
            ):
                if row_index > 20:
                    break

                values = [
                    value if value is not None else ""
                    for value in row
                ]

                if any(value != "" for value in values):
                    rows_contexto.append(values)

        except Exception:
            rows_contexto = []

        banco_textual = detectar_banco(
            worksheet.title,
            rows_contexto
        )

        if not _buscar_banco_catalogo_en_texto(
            banco_textual
        ):
            pendientes.append(
                worksheet.title
            )

    return pendientes


def detectar_bancos_desde_imagenes_excel(
    input_path,
    hojas_objetivo=None,
    max_images_per_sheet=OCR_MAX_IMAGES_PER_SHEET,
    max_seconds=OCR_MAX_SECONDS_PER_FILE
):
    """
    Abre una segunda copia NO read_only del Excel para acceder a ws._images.
    Cada imagen queda asociada a su hoja, por lo que un libro con varios bancos
    puede reconocer un logo diferente en cada pestaña.

    Retorna:
        {
            "Report1": "BAC",
            "Estado": "BANCO AGRÍCOLA"
        }
    """
    if not input_path:
        return {}

    inicio_ocr = time.monotonic()

    engine = _obtener_motor_ocr()
    if engine is None:
        return {}

    objetivo = set(
        hojas_objetivo or []
    )

    encontrados = {}
    workbook_images = None

    try:
        workbook_images = load_workbook(
            input_path,
            data_only=True,
            read_only=False
        )

        for worksheet in workbook_images.worksheets:
            if (time.monotonic() - inicio_ocr) >= max_seconds:
                print(
                    "OCR bancario detenido por límite de tiempo:",
                    f"{max_seconds}s"
                )
                break

            if objetivo and worksheet.title not in objetivo:
                continue

            images = list(
                getattr(
                    worksheet,
                    "_images",
                    []
                )
                or []
            )

            if not images:
                continue

            for image in images[:max_images_per_sheet]:
                if (time.monotonic() - inicio_ocr) >= max_seconds:
                    print(
                        "OCR bancario detenido por límite de tiempo:",
                        f"{max_seconds}s"
                    )
                    break

                try:
                    image_bytes = image._data()
                except Exception as error:
                    print(
                        "No se pudo extraer una imagen de la hoja",
                        worksheet.title,
                        ":",
                        str(error)
                    )
                    continue

                texto_ocr = _ocr_texto_imagen_bytes(
                    image_bytes
                )

                banco = _buscar_banco_catalogo_en_texto(
                    texto_ocr
                )

                if banco:
                    encontrados[
                        worksheet.title
                    ] = banco

                    print(
                        "Banco detectado por imagen/OCR:",
                        worksheet.title,
                        "=>",
                        banco,
                        "| OCR:",
                        texto_ocr[:180]
                    )
                    break

    except Exception as error:
        print(
            "No se pudieron analizar imágenes del Excel:",
            str(error)
        )

    finally:
        if workbook_images is not None:
            try:
                workbook_images.close()
            except Exception:
                pass

    return encontrados



# ============================================================
# CLASIFICACIÓN DE PAÍS: GUATEMALA / EL SALVADOR
# ============================================================

def normalizar_pais_bancario(value):
    texto = clean_text(value)

    if texto in {
        "guatemala",
        "gt",
        "gua",
        "guate"
    }:
        return "GUATEMALA"

    if texto in {
        "el salvador",
        "salvador",
        "sv",
        "slv"
    }:
        return "EL_SALVADOR"

    if "guatemala" in texto:
        return "GUATEMALA"

    if "el salvador" in texto or "salvadoreno" in texto:
        return "EL_SALVADOR"

    return "NO_IDENTIFICADO"


def detectar_pais_bancario(
    sheet_name="",
    rows=None,
    banco="",
    moneda="N/D",
    nombre_archivo=""
):
    """
    Clasifica una hoja bancaria como Guatemala o El Salvador.

    Prioridad:
    1. País escrito explícitamente en la hoja.
    2. Bancos exclusivos o fuertemente asociados a un país.
    3. Moneda de la hoja como respaldo para bancos regionales.
    4. Nombre del archivo como último respaldo.

    No convierte importes ni cambia la moneda detectada.
    """

    rows = rows or []

    contexto_local = [
        str(sheet_name or ""),
        str(banco or "")
    ]

    for row in rows[:40]:
        for value in row:
            if value not in (None, ""):
                contexto_local.append(str(value))

    texto_local = clean_text(
        " | ".join(contexto_local)
    )

    # País explícito dentro de la propia hoja.
    if (
        "el salvador" in texto_local
        or "salvadoreno" in texto_local
        or "salvadorena" in texto_local
    ):
        return "EL_SALVADOR"

    if (
        "guatemala" in texto_local
        or "guatemalteco" in texto_local
        or "guatemalteca" in texto_local
    ):
        return "GUATEMALA"

    banco_texto = clean_text(
        banco
    )

    # Bancos que identifican con fuerza Guatemala.
    senales_gt = (
        "g&t",
        "gyt",
        "banrural",
        "desarrollo rural",
        "agricola mercantil",
        "bantrab",
        "banco de los trabajadores",
        "credito hipotecario nacional",
        "banco inmobiliario",
        "interbanco",
        "banco internacional",
        "vivibanco",
        "banco de antigua",
        "banco azteca",
        "banco inv",
        "ficohsa",
    )

    if any(
        signal in banco_texto
        for signal in senales_gt
    ):
        return "GUATEMALA"

    # Bancos que identifican con fuerza El Salvador.
    senales_sv = (
        "banco agricola",
        "cuscatlan",
        "davivienda",
        "banco hipotecario",
        "banco azul",
        "abank",
        "banco atlantida",
        "apoyo integral",
        "banco integral",
        "fomento agropecuario",
        "citibank",
    )

    if any(
        signal in banco_texto
        for signal in senales_sv
    ):
        return "EL_SALVADOR"

    # Banco Industrial, BAC y Promerica operan/regionalizan nombres.
    # Para ellos la moneda de la hoja es el respaldo más seguro disponible
    # cuando el país no está escrito explícitamente.
    moneda_codigo = normalizar_codigo_moneda(
        moneda
    )

    if moneda_codigo == "GTQ":
        return "GUATEMALA"

    if moneda_codigo in {
        "USD",
        "SVC"
    }:
        return "EL_SALVADOR"

    # Banco Industrial sin otra señal conserva Guatemala como respaldo
    # histórico del parser. Si es El Salvador, normalmente la hoja indicará
    # USD o "El Salvador" y se clasificará antes de llegar aquí.
    if "banco industrial" in banco_texto:
        return "GUATEMALA"

    # Último respaldo: nombre general del archivo.
    archivo = clean_text(
        nombre_archivo
    )

    if "el salvador" in archivo:
        return "EL_SALVADOR"

    if "guatemala" in archivo:
        return "GUATEMALA"

    return "NO_IDENTIFICADO"


def moneda_pais_dashboard(pais, monedas=None):
    """
    Devuelve la moneda visual de la vista país.
    No realiza conversiones.
    """

    pais = normalizar_pais_bancario(
        pais
    )

    monedas = {
        normalizar_codigo_moneda(m)
        for m in (monedas or [])
        if normalizar_codigo_moneda(m)
        not in {
            "",
            "N/D",
            "MULTI"
        }
    }

    if len(monedas) == 1:
        return next(iter(monedas))

    # Para el alcance actual del dashboard:
    # Guatemala se presenta en GTQ y El Salvador en USD.
    if pais == "GUATEMALA":
        return "GTQ"

    if pais == "EL_SALVADOR":
        return "USD"

    if len(monedas) > 1:
        return "MULTI"

    return "N/D"


def _clave_banco_pais_dashboard(
    pais,
    clave_banco
):
    pais = normalizar_pais_bancario(
        pais
    )

    clave_banco = str(
        clave_banco
        or "BANCO NO IDENTIFICADO"
    ).strip()

    return (
        pais
        + "|"
        + clave_banco
    )


def _account_candidate(value):
    if value in (None, ""):
        return ""

    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    # Evitar saldos y fechas.
    if re.search(r"[.,]\d{1,2}$", text):
        return ""
    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text):
        return ""

    digits = re.sub(r"[^0-9]", "", text)
    if 5 <= len(digits) <= 30:
        return digits
    return ""


def detectar_cuenta(sheet_name, rows, header_index=None):
    limit = min((header_index + 1) if header_index is not None else 25, len(rows))

    # A) Formato etiqueta/valor en la misma fila o celda.
    regexes = [
        # Cuenta estándar:
        # Cuenta: 95510021962
        r"(?:numero|nro|no\.?|#)?\s*(?:de\s+)?[#:]?\s*cuenta\s*[:#\-]?\s*([0-9][0-9\- ]{3,30})",

        # Davivienda / estados con etiqueta:
        # Cuenta corriente 95510021962
        r"cuenta\s+(?:corriente|ahorro|monetaria|corriente\s+no\.?)\s*[:#\-]?\s*([0-9][0-9\- ]{3,30})",

        r"account(?:\s+number|\s+no\.?)?\s*[:#\-]?\s*([0-9][0-9\- ]{3,30})",
        r"iban\s*[:#\-]?\s*([A-Z0-9][A-Z0-9\- ]{5,34})",
    ]

    for row in rows[:limit]:
        joined = " | ".join(str(v) for v in row if v not in (None, ""))
        normalized = clean_text(joined)
        for pattern in regexes:
            match = re.search(pattern, normalized, re.I)
            if match:
                digits = re.sub(r"[^0-9]", "", match.group(1))
                if len(digits) >= 5:
                    return digits


    # B) Encabezado tipo tabla: Cuenta / Producto / Account en una fila y valor
    # debajo (muy común en exports de BAC y otros bancos regionales).
    for r in range(max(0, limit - 1)):
        row = rows[r]
        next_row = rows[r + 1] if r + 1 < len(rows) else []
        for c, value in enumerate(row):
            label = compact_text(value)
            if not label:
                continue
            if (
                "cuenta" in label
                or label in {"account", "account number", "producto", "product"}
                or "numero de cuenta" in label
            ):
                # Misma fila: algunos exports dejan una celda vacía entre
                # etiqueta y valor por celdas combinadas. Revisar las próximas 3.
                for cc in range(c + 1, min(len(row), c + 4)):
                    candidate = _account_candidate(row[cc])
                    if candidate:
                        return candidate
                # Fila siguiente: misma columna y vecinas cercanas.
                for cc in range(c, min(len(next_row), c + 3)):
                    candidate = _account_candidate(next_row[cc])
                    if candidate:
                        return candidate

    # C) Nombre de hoja. Solo como último recurso.
    candidates = re.findall(r"(?<!\d)(\d{3,30})(?!\d)", str(sheet_name))
    # Ignora secuencias que parecen periodo 0826 / 2026.
    candidates = [x for x in candidates if x not in {"2024", "2025", "2026", "2027", "2028"}]
    if candidates:
        return max(candidates, key=len)

    return ""


def detectar_columnas(rows):
    best = None

    for row_index, row in enumerate(rows):
        columnas = {}
        confidences = {}

        for index, value in enumerate(row):
            field, score = _header_field(value)
            if not field:
                continue

            # "valor" sin contexto puede ser monto, pero no debe desplazar
            # débito/crédito/saldo si ya existe un match más fuerte.
            if field not in columnas or score > confidences.get(field, 0):
                columnas[field] = index
                confidences[field] = score

        has_date = "fecha" in columnas
        has_desc_or_ref = "descripcion" in columnas or "referencia" in columnas
        has_pair = "debito" in columnas and "credito" in columnas
        has_amount = "monto" in columnas
        has_financial = has_pair or has_amount or (
            "saldo" in columnas and ("debito" in columnas or "credito" in columnas)
        )

        if not (has_date and has_desc_or_ref and has_financial):
            continue

        score = 0
        weights = {
            "fecha": 4, "descripcion": 3, "referencia": 2, "codigo": 1,
            "debito": 3, "credito": 3, "monto": 3, "saldo": 3, "tipo": 1,
        }
        for field in columnas:
            score += weights.get(field, 1)
        score += sum(confidences.values())

        candidate = (score, row_index, columnas)
        if best is None or candidate[0] > best[0]:
            best = candidate

    if best:
        return best[1], best[2]
    return None, {}


def get_column(row, columns, name):
    index = columns.get(name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def _looks_like_summary(row):
    values = [str(v) for v in row if v not in (None, "")]
    if not values:
        return True
    joined = clean_text(" | ".join(values))
    first = compact_text(values[0])
    first_words = re.sub(r"[^a-z0-9 ]+", " ", first)
    first_words = re.sub(r"\s+", " ", first_words).strip()

    normalized_prefixes = [
        re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]+", " ", compact_text(p))).strip()
        for p in SUMMARY_PREFIXES
    ]
    if any(first_words.startswith(prefix) for prefix in normalized_prefixes if prefix):
        return True
    if (
        first_words.startswith("no credito")
        or first_words.startswith("no debito")
        or first_words.startswith("cantidad credito")
        or first_words.startswith("cantidad debito")
        or first_words.startswith("monto credito")
        or first_words.startswith("monto debito")
    ):
        return True
    if "resumen de estado" in joined or "summary of" in joined:
        return True
    return False


def _movement_direction(text):
    t = compact_text(text)
    if not t:
        return None

    # Match de palabras completas cuando son códigos cortos.
    tokens = set(re.findall(r"[a-z]+", t))
    if tokens.intersection({"cr", "credit", "credito", "abono", "deposito", "deposit", "haber", "ingreso"}):
        return "credit"
    if tokens.intersection({"dr", "db", "debit", "debito", "cargo", "retiro", "debe", "egreso"}):
        return "debit"

    if any(h in t for h in CREDIT_HINTS if len(h) > 2):
        return "credit"
    if any(h in t for h in DEBIT_HINTS if len(h) > 2):
        return "debit"
    return None


def detectar_saldo_inicial(rows, header_index):
    limit = min(header_index + 1, len(rows)) if header_index is not None else min(25, len(rows))

    for r in range(limit):
        row = rows[r]
        for c, value in enumerate(row):
            label = compact_text(value)
            if not label:
                continue

            if any(key in label for key in (
                "saldo inicial", "saldo anterior", "balance inicial",
                "beginning balance", "opening balance", "previous balance",
            )):
                # Número embebido en la misma celda.
                numbers = re.findall(r"[-+]?\d[\d., ]*\d|[-+]?\d", str(value))
                for n in reversed(numbers):
                    parsed = parse_number(n)
                    if parsed is not None:
                        return parsed

                # Celda siguiente.
                if c + 1 < len(row):
                    parsed = parse_number(row[c + 1])
                    if parsed is not None:
                        return parsed

                # Misma columna, fila siguiente.
                if r + 1 < len(rows) and c < len(rows[r + 1]):
                    parsed = parse_number(rows[r + 1][c])
                    if parsed is not None:
                        return parsed

    return None


def extraer_transaccion(
    row,
    columns,
    banco,
    cuenta,
    expected_month=None,
    expected_year=None,
    previous_balance=None,
    previous_date=None,
    saldo_inicial_cuenta=None,
):
    raw_date = get_column(row, columns, "fecha")
    referencia = get_column(row, columns, "referencia")
    codigo = get_column(row, columns, "codigo")
    descripcion = get_column(row, columns, "descripcion")
    tipo = get_column(row, columns, "tipo")

    fecha = clean_date(raw_date, expected_month, expected_year)

    # Permite estados donde la fecha aparece una sola vez y las siguientes
    # líneas pertenecen al mismo día, pero evita usar esto en resúmenes.
    if not isinstance(fecha, datetime) and previous_date and raw_date in (None, ""):
        fecha = previous_date

    if not isinstance(fecha, datetime):
        return None

    if _looks_like_summary(row):
        return None

    debit_raw = get_column(row, columns, "debito")
    credit_raw = get_column(row, columns, "credito")
    amount_raw = get_column(row, columns, "monto")
    balance_raw = get_column(row, columns, "saldo")

    debit_val = parse_number(debit_raw)
    credit_val = parse_number(credit_raw)
    amount_val = parse_number(amount_raw)
    balance_val = parse_number(balance_raw)

    debito = -abs(debit_val) if debit_val not in (None, 0) else 0.0
    credito = abs(credit_val) if credit_val not in (None, 0) else 0.0

    # Formato de una sola columna de monto.
    if debito == 0 and credito == 0 and amount_val not in (None, 0):
        direction_text = " ".join(str(v) for v in (tipo, codigo, descripcion) if v not in (None, ""))
        direction = _movement_direction(direction_text)

        if amount_val < 0:
            direction = "debit"

        # El saldo corrido permite resolver formatos que muestran el monto
        # siempre positivo pero no indican explícitamente D/C.
        if direction is None and previous_balance is not None and balance_val is not None:
            delta = balance_val - previous_balance
            if abs(delta) > 0.000001:
                direction = "credit" if delta > 0 else "debit"

        if direction == "debit":
            debito = -abs(amount_val)
        else:
            # Para columnas firmadas, un positivo normalmente es crédito.
            credito = abs(amount_val)

    # Formato sin monto explícito, pero con saldo corrido.
    if debito == 0 and credito == 0 and balance_val is not None and previous_balance is not None:
        delta = balance_val - previous_balance
        if abs(delta) > 0.000001:
            if delta > 0:
                credito = abs(delta)
            else:
                debito = -abs(delta)

    texto_desc = str(descripcion).strip() if descripcion not in (None, "") else ""
    texto_ref = str(referencia).strip() if referencia not in (None, "") else ""
    texto_codigo = str(codigo).strip() if codigo not in (None, "") else ""

    # Una transacción real necesita monto/movimiento y alguna identificación.
    has_money = debito != 0 or credito != 0 or balance_val is not None
    has_identity = bool(texto_desc or texto_ref or texto_codigo)
    if not (has_money and has_identity):
        return None

    return {
        "banco": valor_o_nd(banco),
        "cuenta": valor_o_nd(cuenta),
        "fecha": fecha,
        "referencia": texto_ref or texto_codigo,
        "codigo": texto_codigo,
        "descripcion": texto_desc or texto_codigo or texto_ref,
        "debito": debito,
        "credito": credito,
        "saldo": balance_val if balance_val is not None else 0.0,
        "saldo_inicial_cuenta": saldo_inicial_cuenta,
        # Conserva si el dato existía realmente en el estado de cuenta.
        "_tiene_debito": debit_val is not None or (amount_val not in (None, 0) and debito != 0),
        "_tiene_credito": credit_val is not None or (amount_val not in (None, 0) and credito != 0),
        "_tiene_saldo": balance_val is not None,
    }


def procesar_hoja(
    worksheet,
    moneda_respaldo="N/D",
    nombre_archivo="",
    banco_respaldo_ocr=""
):
    rows = []

    for row in worksheet.iter_rows(values_only=True):
        values = [value if value is not None else "" for value in row]
        if any(value != "" for value in values):
            rows.append(values)

    if not rows:
        return []

    header_index, columns = detectar_columnas(rows)
    if header_index is None:
        return []

    # PRIORIDAD DE IDENTIFICACIÓN BANCARIA:
    # 1) Nombre del archivo (rápido y confiable para los estados SV).
    # 2) Nombre de hoja / texto de celdas.
    # 3) OCR de imágenes únicamente como último respaldo.
    banco_archivo = detectar_banco_por_nombre_archivo(
        nombre_archivo
    )

    if banco_archivo:
        banco = banco_archivo
    else:
        banco = detectar_banco(worksheet.title, rows)

        banco_catalogo = _buscar_banco_catalogo_en_texto(
            banco
        )

        if banco_catalogo:
            banco = banco_catalogo
        elif banco_respaldo_ocr:
            banco = banco_respaldo_ocr

    cuenta = detectar_cuenta(worksheet.title, rows, header_index)

    # --------------------------------------------------------
    # MONEDA
    # --------------------------------------------------------
    # 1) Prioridad máxima: moneda escrita en el estado de cuenta.
    # 2) Si no existe, usar la moneda contextual del archivo.
    # 3) Solo al final revisar el number_format de las celdas.
    #
    # Esto evita que un formato viejo de Excel (ej. Q) gane sobre
    # un texto real como "MONETARIO (USD)".
    try:
        moneda_textual = detectar_moneda_textual(
            rows,
            worksheet.title,
            nombre_archivo
        )

        if moneda_textual not in {"N/D", "MULTI"}:
            moneda = moneda_textual

        elif moneda_respaldo not in {"N/D", "MULTI", ""}:
            moneda = moneda_respaldo

        else:
            moneda = detectar_moneda(
                worksheet,
                rows,
                header_index
            )

    except Exception as currency_error:
        print(
            "No se pudo detectar moneda en hoja",
            worksheet.title,
            ":",
            str(currency_error)
        )

        moneda = (
            moneda_respaldo
            if moneda_respaldo not in {"N/D", "MULTI", ""}
            else "N/D"
        )

    # --------------------------------------------------------
    # PAÍS DE LA CUENTA / HOJA
    # --------------------------------------------------------
    pais = detectar_pais_bancario(
        sheet_name=worksheet.title,
        rows=rows,
        banco=banco,
        moneda=moneda,
        nombre_archivo=nombre_archivo
    )

    expected_month, expected_year = detectar_periodo(rows, worksheet.title)
    saldo_inicial = detectar_saldo_inicial(rows, header_index)

    transactions = []
    previous_balance = saldo_inicial
    previous_date = None
    invalid_streak = 0

    for row in rows[header_index + 1:]:
        transaction = extraer_transaccion(
            row,
            columns,
            banco,
            cuenta,
            expected_month=expected_month,
            expected_year=expected_year,
            previous_balance=previous_balance,
            previous_date=previous_date,
            saldo_inicial_cuenta=saldo_inicial,
        )

        if transaction:
            # Metadatos visuales; no modifican ningún importe.
            transaction["moneda"] = moneda
            transaction["pais"] = pais
            transactions.append(transaction)
            previous_date = transaction["fecha"]
            if transaction["saldo"] != 0 or get_column(row, columns, "saldo") not in (None, ""):
                previous_balance = transaction["saldo"]
            invalid_streak = 0
        else:
            invalid_streak += 1

            # No corta inmediatamente al ver un resumen: algunos bancos ponen
            # subtotales entre grupos. Solo deja de recorrer después de un bloque
            # largo sin movimientos reales.
            if transactions and invalid_streak >= 25:
                break

    return transactions


# ============================================================
# COPIAR ESTILO DE FILA
# ============================================================

def copiar_estilo_fila(
    worksheet,
    source_row,
    target_row
):

    for column in range(
        1,
        worksheet.max_column + 1
    ):

        source = worksheet.cell(
            row=source_row,
            column=column
        )

        target = worksheet.cell(
            row=target_row,
            column=column
        )

        # No tocar MergedCell

        if (
            es_celda_combinada(source)
            or es_celda_combinada(target)
        ):
            continue

        if source.has_style:

            target._style = copy(
                source._style
            )

        if source.number_format:

            target.number_format = (
                source.number_format
            )

        if source.alignment:

            target.alignment = copy(
                source.alignment
            )

        if source.protection:

            target.protection = copy(
                source.protection
            )


# ============================================================
# LIMPIAR CONTENIDO
# ============================================================

def limpiar_filas(
    worksheet,
    start_row,
    end_row,
    start_column=1,
    end_column=None
):

    if end_column is None:

        end_column = worksheet.max_column

    for row in range(
        start_row,
        end_row + 1
    ):

        for column in range(
            start_column,
            end_column + 1
        ):

            limpiar_celda_segura(
                worksheet,
                row,
                column
            )


# ============================================================
# ESCRIBIR ESTADOS CONSOLIDADOS
# ============================================================

def escribir_estados_consolidados(
    workbook,
    transactions,
    sheet_name=None
):

    target_sheet = sheet_name or SHEET_ESTADOS

    if target_sheet not in workbook.sheetnames:

        raise Exception(
            f"No existe la hoja '{target_sheet}'"
        )

    ws = workbook[
        target_sheet
    ]

    start_row = 2

    limpiar_filas(
        ws,
        start_row,
        ws.max_row,
        1,
        8
    )

    required_last_row = (
        start_row
        + len(transactions)
        - 1
    )

    if required_last_row > ws.max_row:

        old_max_row = ws.max_row

        for row in range(
            old_max_row + 1,
            required_last_row + 1
        ):

            copiar_estilo_fila(
                ws,
                start_row,
                row
            )

            ws.row_dimensions[
                row
            ].height = ws.row_dimensions[
                start_row
            ].height

    for index, transaction in enumerate(
        transactions,
        start=start_row
    ):

        escribir_celda_segura(
            ws,
            index,
            1,
            valor_o_nd(transaction["banco"])
        )

        escribir_celda_segura(
            ws,
            index,
            2,
            valor_o_nd(transaction["cuenta"])
        )

        cell = obtener_celda_segura(
            ws,
            index,
            3
        )

        if cell:

            cell.value = transaction["fecha"]

            cell.number_format = "dd/mm/yyyy"

        escribir_celda_segura(
            ws,
            index,
            4,
            valor_o_nd(transaction["referencia"])
        )

        escribir_celda_segura(
            ws,
            index,
            5,
            valor_o_nd(transaction["descripcion"])
        )

        escribir_celda_segura(
            ws,
            index,
            6,
            transaction["debito"] if transaction.get("_tiene_debito") else 0.00
        )

        escribir_celda_segura(
            ws,
            index,
            7,
            transaction["credito"] if transaction.get("_tiene_credito") else 0.00
        )

        escribir_celda_segura(
            ws,
            index,
            8,
            transaction["saldo"] if transaction.get("_tiene_saldo") else "N/D"
        )


# ============================================================
# NORMALIZAR NOMBRE BANCO
# ============================================================

def banco_corto(banco):

    texto = clean_text(banco)

    if "agricola" in texto:
        return "AGRÍCOLA"

    if "cuscatlan" in texto:
        return "CUSCATLÁN"

    if "bac" in texto:
        return "BAC"

    if "promerica" in texto:
        return "PROMERICA"

    if "davivienda" in texto:
        return "DAVIVIENDA"

    if "industrial" in texto:
        return "INDUSTRIAL"

    if "g&t" in texto or "gyt" in texto:
        return "G&T"

    if "banrural" in texto or texto == "ban":
        return "BANRURAL"

    return str(banco).strip()


# ============================================================
# ESCRIBIR SALDOS POR CUENTA
# ============================================================

def escribir_saldos_por_cuenta(
    workbook,
    transactions,
    sheet_name=None
):
    target_sheet = sheet_name or SHEET_SALDOS

    if target_sheet not in workbook.sheetnames:
        raise Exception(f"No existe la hoja '{target_sheet}'")

    ws = workbook[target_sheet]
    start_row = 5
    cuentas = {}

    for transaction in transactions:
        clave = (transaction["banco"], transaction["cuenta"])
        cuentas.setdefault(clave, []).append(transaction)

    limpiar_filas(ws, start_row, ws.max_row, 1, 4)

    required_last_row = start_row + len(cuentas) - 1
    if required_last_row > ws.max_row:
        old_max_row = ws.max_row
        for row in range(old_max_row + 1, required_last_row + 1):
            copiar_estilo_fila(ws, start_row, row)

    row_number = start_row

    for (banco, cuenta), movimientos in cuentas.items():
        movimientos = sorted(
            movimientos,
            key=lambda x: x["fecha"] if isinstance(x["fecha"], datetime) else datetime.min
        )
        if not movimientos:
            continue

        primero = movimientos[0]
        ultimo = movimientos[-1]

        saldo_explicito = primero.get("saldo_inicial_cuenta")
        if saldo_explicito is not None:
            saldo_inicial = saldo_explicito
        elif primero.get("_tiene_saldo"):
            # saldo_nuevo = saldo_anterior + debito + credito
            saldo_inicial = primero["saldo"] - primero["debito"] - primero["credito"]
        else:
            saldo_inicial = "N/D"

        saldo_final = ultimo["saldo"] if ultimo.get("_tiene_saldo") else "N/D"

        escribir_celda_segura(ws, row_number, 1, valor_o_nd(banco_corto(banco)))
        escribir_celda_segura(ws, row_number, 2, valor_o_nd(cuenta))
        escribir_celda_segura(ws, row_number, 3, saldo_inicial)
        escribir_celda_segura(ws, row_number, 4, saldo_final)
        row_number += 1


# ============================================================
# ESCRIBIR REPORTE CREDITOS DIARIOS
# ============================================================

def escribir_reporte_creditos(
    workbook,
    transactions,
    sheet_name=None
):
    """
    Construye el REPORTE CREDITOS DIARIOS sin encoger la plantilla.

    Regla importante:
    - La plantilla define un ancho mínimo de columnas de cuentas.
    - Si el archivo trae menos cuentas, se conservan columnas finales de la
      plantilla para que TOTAL CRÉDITOS y CUENTAS CON ABONO no se desplacen
      hacia la izquierda ni hereden formatos de moneda incorrectos.
    - Si el archivo trae más cuentas que la plantilla, el reporte se expande
      y los dos campos de resumen se mueven al final con su formato correcto.
    """

    target_sheet = sheet_name or SHEET_REPORTE

    if target_sheet not in workbook.sheetnames:
        raise Exception(f"No existe la hoja '{target_sheet}'")

    ws = workbook[target_sheet]

    data = {}
    cuentas = {}

    for transaction in transactions:
        fecha = transaction["fecha"]

        if not isinstance(fecha, (datetime, date)):
            continue

        fecha_key = fecha.date() if isinstance(fecha, datetime) else fecha
        banco = banco_corto(transaction["banco"])
        cuenta = str(transaction["cuenta"])
        clave = (banco, cuenta)

        cuentas[clave] = True
        data.setdefault(fecha_key, {}).setdefault(banco, {}).setdefault(cuenta, 0.0)
        data[fecha_key][banco][cuenta] += transaction["credito"]

    cuentas_ordenadas = sorted(cuentas.keys(), key=lambda x: (x[0], x[1]))

    header_bank_row = 3
    header_account_row = 4
    start_row = 5

    # --------------------------------------------------------
    # 1) LEER LA ESTRUCTURA ORIGINAL DE LA PLANTILLA
    #    antes de limpiar cualquier celda.
    # --------------------------------------------------------

    template_total_col = None
    template_count_col = None

    for col in range(2, ws.max_column + 1):
        header = clean_text(ws.cell(row=header_account_row, column=col).value)

        if template_total_col is None and "total" in header and (
            "credito" in header or "credit" in header
        ):
            template_total_col = col

        if template_count_col is None and "cuentas" in header and "abono" in header:
            template_count_col = col

    # Fallback defensivo para plantillas equivalentes.
    if template_total_col is None:
        template_total_col = max(3, ws.max_column - 1)

    if template_count_col is None:
        template_count_col = template_total_col + 1

    template_last_account_col = max(2, template_total_col - 1)
    template_account_slots = max(0, template_total_col - 2)

    # Guardar el banco de cada columna de cuenta. En encabezados combinados,
    # openpyxl solo conserva el valor en la primera celda; propagamos ese banco
    # a las columnas siguientes del mismo bloque.
    template_accounts = []
    banco_actual = ""

    for col in range(2, template_total_col):
        banco_cell = ws.cell(row=header_bank_row, column=col)
        banco_valor = banco_cell.value

        if banco_valor not in (None, ""):
            banco_actual = str(banco_valor).strip()

        cuenta_valor = ws.cell(row=header_account_row, column=col).value

        if cuenta_valor not in (None, ""):
            template_accounts.append((banco_actual, str(cuenta_valor).strip()))

    # --------------------------------------------------------
    # 2) DEFINIR LAS CUENTAS QUE SE MOSTRARÁN
    # --------------------------------------------------------

    # El archivo de entrada manda en las primeras columnas.
    display_accounts = list(cuentas_ordenadas)

    # Si trae menos cuentas que el diseño base, NO encogemos la hoja.
    # Conservamos el ancho visual, pero las columnas de relleno quedan vacías.
    # Nunca reutilizamos bancos/cuentas antiguos de la plantilla.
    if len(display_accounts) < template_account_slots:
        needed = template_account_slots - len(display_accounts)
        display_accounts.extend([
            ("", "")
            for _ in range(needed)
        ])

    # Si vienen más cuentas que la plantilla, simplemente expandimos.
    total_column = 2 + len(display_accounts)
    cuentas_abono_column = total_column + 1

    # --------------------------------------------------------
    # 3) PREPARAR ESTILOS CUANDO EL REPORTE SE EXPANDE
    # --------------------------------------------------------

    # Copiar estilo de una columna a otra, fila por fila.
    def copiar_estilo_columna(source_col, target_col, last_row):
        if source_col == target_col:
            return

        for row in range(1, last_row + 1):
            source = ws.cell(row=row, column=source_col)
            target = ws.cell(row=row, column=target_col)

            if es_celda_combinada(source) or es_celda_combinada(target):
                continue

            if source.has_style:
                target._style = copy(source._style)

            if source.number_format:
                target.number_format = source.number_format

            if source.alignment:
                target.alignment = copy(source.alignment)

            if source.protection:
                target.protection = copy(source.protection)

    fechas = sorted(data.keys())

    if not fechas:
        return

    required_last_row = start_row + len(fechas) - 1
    style_last_row = max(ws.max_row, required_last_row)

    # Si las cuentas ocupan las columnas donde antes estaban TOTAL/CUENTAS,
    # convertir esas columnas a estilo de cuenta.
    if total_column > template_total_col:
        for col in range(template_total_col, total_column):
            copiar_estilo_columna(template_last_account_col, col, style_last_row)

        # Llevar los estilos originales de TOTAL y CUENTAS al nuevo extremo.
        copiar_estilo_columna(template_total_col, total_column, style_last_row)
        copiar_estilo_columna(template_count_col, cuentas_abono_column, style_last_row)

    # --------------------------------------------------------
    # 4) DESCOMBINAR SOLO EL ENCABEZADO DINÁMICO
    # --------------------------------------------------------

    merges_to_remove = []

    for merged_range in list(ws.merged_cells.ranges):
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row

        if (
            min_row <= header_account_row
            and max_row >= header_bank_row
            and max_col >= 2
        ):
            merges_to_remove.append(str(merged_range))

    for merged_range in merges_to_remove:
        try:
            ws.unmerge_cells(merged_range)
        except Exception:
            pass

    # --------------------------------------------------------
    # 5) LIMPIAR VALORES, NO ESTILOS
    # --------------------------------------------------------

    clear_to_col = max(ws.max_column, cuentas_abono_column)

    limpiar_filas(
        ws,
        header_bank_row,
        max(ws.max_row, required_last_row),
        1,
        clear_to_col
    )

    # Restaurar encabezados base.
    escribir_celda_segura(ws, header_bank_row, 1, "BANCO")
    escribir_celda_segura(ws, header_account_row, 1, "Fecha")

    # --------------------------------------------------------
    # 6) CREAR ENCABEZADOS DE CUENTAS Y AGRUPAR BANCOS
    # --------------------------------------------------------

    column_map = {}

    for offset, (banco, cuenta) in enumerate(display_accounts):
        col = 2 + offset
        cuenta_texto = str(cuenta).strip() if str(cuenta or "").strip() else ""

        escribir_celda_segura(ws, header_account_row, col, cuenta_texto)

        # Solo las cuentas realmente presentes en el archivo fuente entran al
        # mapa de datos. Las columnas conservadas de plantilla quedan vacías.
        incoming_key = (banco, cuenta)
        if incoming_key in cuentas:
            column_map[incoming_key] = col

    # Agrupar encabezados contiguos del mismo banco, igual que la plantilla.
    if display_accounts:
        group_start = 0

        while group_start < len(display_accounts):
            banco = display_accounts[group_start][0]
            group_end = group_start

            while (
                group_end + 1 < len(display_accounts)
                and clean_text(display_accounts[group_end + 1][0]) == clean_text(banco)
            ):
                group_end += 1

            start_col = 2 + group_start
            end_col = 2 + group_end

            escribir_celda_segura(
                ws,
                header_bank_row,
                start_col,
                valor_o_nd(banco) if banco else ""
            )

            if end_col > start_col:
                try:
                    ws.merge_cells(
                        start_row=header_bank_row,
                        start_column=start_col,
                        end_row=header_bank_row,
                        end_column=end_col
                    )
                except Exception:
                    pass

            group_start = group_end + 1

    # Resúmenes SIEMPRE al final del ancho lógico del reporte.
    escribir_celda_segura(
        ws,
        header_account_row,
        total_column,
        "TOTAL CRÉDITOS"
    )

    escribir_celda_segura(
        ws,
        header_account_row,
        cuentas_abono_column,
        "CUENTAS CON ABONO"
    )

    # --------------------------------------------------------
    # 7) ASEGURAR FILAS SUFICIENTES
    # --------------------------------------------------------

    if required_last_row > ws.max_row:
        old_max_row = ws.max_row

        for row in range(old_max_row + 1, required_last_row + 1):
            copiar_estilo_fila(ws, start_row, row)

    # --------------------------------------------------------
    # 8) ESCRIBIR DATOS
    # --------------------------------------------------------

    for index, fecha in enumerate(fechas, start=start_row):
        fecha_cell = obtener_celda_segura(ws, index, 1)

        if fecha_cell:
            fecha_cell.value = datetime(fecha.year, fecha.month, fecha.day)
            fecha_cell.number_format = "dd/mm/yyyy"

        total = 0.0
        cuentas_con_abono = 0

        # Inicializar columnas de cuenta con 0.00 para mantener valores numéricos.
        # Evita celdas vacías/N/D en el reporte de créditos diarios.
        for col in range(2, total_column):
            escribir_celda_segura(ws, index, col, 0.00)

        # Después escribir valores reales del archivo fuente.
        for clave, column in column_map.items():
            banco, cuenta = clave

            valor = data.get(fecha, {}).get(banco, {}).get(cuenta, 0.0)

            if valor != 0:
                escribir_celda_segura(ws, index, column, valor)
                cuentas_con_abono += 1
                total += valor

        escribir_celda_segura(ws, index, total_column, total)
        escribir_celda_segura(ws, index, cuentas_abono_column, cuentas_con_abono)

        # Forzar formato entero para CUENTAS CON ABONO.
        count_cell = obtener_celda_segura(ws, index, cuentas_abono_column)
        if count_cell:
            count_cell.number_format = "0"


# ============================================================
# GRÁFICAS DINÁMICAS DEL TABLERO
# ============================================================

def ultima_fila_con_valor(worksheet, column, start_row):
    """Devuelve la última fila con contenido real en una columna."""
    last_row = start_row - 1

    for row in range(start_row, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=column).value
        if value not in (None, ""):
            last_row = row

    return last_row


def _obtener_formula_fuente(data_source):
    """Lee la fórmula de origen de categorías/valores de una serie."""
    if data_source is None:
        return ""

    for attr in ("strRef", "numRef", "multiLvlStrRef"):
        ref = getattr(data_source, attr, None)
        if ref is not None:
            formula = getattr(ref, "f", None)
            if formula:
                return str(formula)

    return ""


def _actualizar_formula_fuente(data_source, formula):
    """
    Cambia el rango de una serie existente y elimina el cache viejo
    del gráfico para obligar a Excel/Google Sheets a leer los datos nuevos.
    """
    if data_source is None:
        return False

    for attr in ("strRef", "numRef", "multiLvlStrRef"):
        ref = getattr(data_source, attr, None)
        if ref is None:
            continue

        ref.f = formula

        # Muy importante: no conservar datos cacheados de la plantilla.
        for cache_attr in (
            "strCache",
            "numCache",
            "multiLvlStrCache"
        ):
            if hasattr(ref, cache_attr):
                try:
                    setattr(ref, cache_attr, None)
                except Exception:
                    pass

        return True

    return False


def actualizar_graficas_dinamicas(
    worksheet,
    evolution_end,
    cuenta_end,
    cantidad_cuentas
):
    """
    Adapta las dos gráficas de TABLERO CREDITOS a los datos reales.

    - Evolución: A15:B<última fecha real>
    - Créditos por cuenta: G15:H<última cuenta real>
    - Elimina cache de la plantilla para evitar que aparezcan bancos viejos.
    - Aumenta el ancho de la gráfica de cuentas cuando hay muchas cuentas.
    """
    charts = getattr(worksheet, "_charts", [])

    if not charts:
        return

    sheet_ref = worksheet.title.replace("'", "''")

    evolution_end = max(15, evolution_end)
    cuenta_end = max(15, cuenta_end)
    cantidad_cuentas = max(1, cantidad_cuentas)

    evolution_cat = f"'{sheet_ref}'!$A$15:$A${evolution_end}"
    evolution_val = f"'{sheet_ref}'!$B$15:$B${evolution_end}"

    account_cat = f"'{sheet_ref}'!$G$15:$G${cuenta_end}"
    account_val = f"'{sheet_ref}'!$H$15:$H${cuenta_end}"

    evolution_updated = False
    accounts_updated = False

    for chart in charts:
        series_list = getattr(chart, "series", [])

        for series in series_list:
            current_cat = _obtener_formula_fuente(
                getattr(series, "cat", None)
            )
            current_val = _obtener_formula_fuente(
                getattr(series, "val", None)
            )

            current = f"{current_cat} {current_val}".upper()

            # Gráfica 1: Evolución del crédito seleccionado.
            if "$A$15" in current or "$B$15" in current:
                _actualizar_formula_fuente(
                    getattr(series, "cat", None),
                    evolution_cat
                )
                _actualizar_formula_fuente(
                    getattr(series, "val", None),
                    evolution_val
                )
                evolution_updated = True

            # Gráfica 2: Créditos por cuenta en el rango.
            elif "$G$15" in current or "$H$15" in current:
                _actualizar_formula_fuente(
                    getattr(series, "cat", None),
                    account_cat
                )
                _actualizar_formula_fuente(
                    getattr(series, "val", None),
                    account_val
                )
                accounts_updated = True

                # Con más cuentas, hacer la gráfica más ancha para que
                # no desaparezcan o se amontonen las etiquetas.
                try:
                    if cantidad_cuentas <= 8:
                        chart.width = 15.0
                    else:
                        chart.width = min(
                            32.0,
                            15.0 + ((cantidad_cuentas - 8) * 1.15)
                        )

                    chart.height = 8.0

                    # Reducir el espacio entre columnas cuando hay muchas.
                    if hasattr(chart, "gapWidth"):
                        chart.gapWidth = max(
                            35,
                            int(150 * 8 / cantidad_cuentas)
                        )
                except Exception:
                    pass

    # Fallback por orden de la plantilla si alguna referencia original
    # fue cambiada manualmente y no se pudo identificar por su rango.
    if not evolution_updated and len(charts) >= 1:
        try:
            series = charts[0].series[0]
            _actualizar_formula_fuente(series.cat, evolution_cat)
            _actualizar_formula_fuente(series.val, evolution_val)
        except Exception:
            pass

    if not accounts_updated and len(charts) >= 2:
        try:
            chart = charts[1]
            series = chart.series[0]
            _actualizar_formula_fuente(series.cat, account_cat)
            _actualizar_formula_fuente(series.val, account_val)

            if cantidad_cuentas <= 8:
                chart.width = 15.0
            else:
                chart.width = min(
                    32.0,
                    15.0 + ((cantidad_cuentas - 8) * 1.15)
                )
        except Exception:
            pass


def forzar_recalculo_excel(workbook):
    """Marca el libro para recalcular fórmulas y gráficos al abrirlo."""
    try:
        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.calcMode = "auto"
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
    except Exception:
        pass


# ============================================================
# ACTUALIZAR TABLERO
# ============================================================

def actualizar_tablero(
    workbook,
    transactions,
    sheet_name=None,
    reporte_sheet_name=None,
    saldos_sheet_name=None
):
    """
    Actualiza TABLERO CREDITOS usando exclusivamente las hojas del mismo
    juego/moneda.

    Incluye:
    - Créditos en rango.
    - Promedio diario.
    - Mayor día.
    - Cuentas con abono.
    - Saldo disponible (saldo final real de SALDOS POR CUENTA).
    - Total de créditos por cuenta.

    Reglas de SALDO DISPONIBLE:
    - CUENTA seleccionada: saldo final de esa cuenta.
    - BANCO seleccionado: suma de saldos finales de sus cuentas.
    - TODAS / TODOS: suma consolidada de saldos finales del juego de moneda.
    """
    target_sheet = sheet_name or SHEET_TABLERO
    target_reporte = reporte_sheet_name or SHEET_REPORTE
    target_saldos = saldos_sheet_name or SHEET_SALDOS

    required_sheets = {
        target_sheet,
        target_reporte,
        target_saldos
    }

    if not required_sheets.issubset(set(workbook.sheetnames)):
        return

    ws = workbook[target_sheet]
    reporte = workbook[target_reporte]
    saldos = workbook[target_saldos]

    fechas = [
        t["fecha"] for t in transactions
        if isinstance(t.get("fecha"), (datetime, date))
    ]

    if fechas:
        fechas = sorted(fechas)
        fecha_min, fecha_max = fechas[0], fechas[-1]

        if isinstance(fecha_min, date) and not isinstance(fecha_min, datetime):
            fecha_min = datetime(fecha_min.year, fecha_min.month, fecha_min.day)

        if isinstance(fecha_max, date) and not isinstance(fecha_max, datetime):
            fecha_max = datetime(fecha_max.year, fecha_max.month, fecha_max.day)

        cell = obtener_celda_segura(ws, 5, 1)
        if cell:
            cell.value = fecha_min
            cell.number_format = "dd/mm/yyyy"

        cell = obtener_celda_segura(ws, 5, 3)
        if cell:
            cell.value = fecha_max
            cell.number_format = "dd/mm/yyyy"

    escribir_celda_segura(ws, 5, 5, "TODAS")
    escribir_celda_segura(ws, 5, 7, "TODOS")

    total_col = None
    count_col = None

    for col in range(2, reporte.max_column + 1):
        header = clean_text(reporte.cell(row=4, column=col).value)

        if "total" in header and ("credito" in header or "credit" in header):
            total_col = col

        if "cuentas" in header and "abono" in header:
            count_col = col

    if total_col is None:
        total_col = max(2, reporte.max_column - 1)

    if count_col is None:
        count_col = total_col + 1

    last_account_col = max(2, total_col - 1)
    first_account_letter = get_column_letter(2)
    last_account_letter = get_column_letter(last_account_col)
    total_letter = get_column_letter(total_col)
    count_letter = get_column_letter(count_col)

    report_start = 5
    report_end = ultima_fila_con_valor(reporte, 1, report_start)

    if report_end < report_start:
        return

    tablero_start = 15

    for row in range(tablero_start, max(ws.max_row, tablero_start + 100) + 1):
        for col in range(1, 5):
            limpiar_celda_segura(ws, row, col)

    max_evolution_rows = max(19, report_end - report_start + 1)
    required_end = tablero_start + max_evolution_rows - 1

    if required_end > ws.max_row:
        old_max = ws.max_row
        for row in range(old_max + 1, required_end + 1):
            copiar_estilo_fila(ws, 15, row)

    for i in range(max_evolution_rows):
        row = tablero_start + i
        report_row = report_start + i

        if report_row > report_end:
            continue

        formula_fecha = (
            f'=IF(AND(\'{target_reporte}\'!A{report_row}>=$A$5,'
            f'\'{target_reporte}\'!A{report_row}<=$C$5),'
            f'\'{target_reporte}\'!A{report_row},"")'
        )
        escribir_celda_segura(ws, row, 1, formula_fecha)

        formula_credito = (
            f'=IF(A{row}="",0,'
            f'IF($E$5<>"TODAS",'
            f'IFERROR(INDEX(\'{target_reporte}\'!${first_account_letter}$5:${last_account_letter}${report_end},'
            f'MATCH(A{row},\'{target_reporte}\'!$A$5:$A${report_end},0),'
            f'MATCH($E$5,\'{target_reporte}\'!${first_account_letter}$4:${last_account_letter}$4,0)),0),'
            f'IF($G$5="TODOS",'
            f'IFERROR(SUMIF(\'{target_reporte}\'!$A$5:$A${report_end},A{row},'
            f'\'{target_reporte}\'!${total_letter}$5:${total_letter}${report_end}),0),0)))'
        )
        escribir_celda_segura(ws, row, 2, formula_credito)

        if i == 0:
            escribir_celda_segura(ws, row, 3, 0)
        else:
            escribir_celda_segura(ws, row, 3, f'=B{row}-B{row - 1}')

        formula_cuentas = (
            f'=IF(A{row}="",0,'
            f'IF($G$5="TODOS",IFERROR(INDEX('
            f'\'{target_reporte}\'!${count_letter}$5:${count_letter}${report_end},'
            f'MATCH(A{row},\'{target_reporte}\'!$A$5:$A${report_end},0)),0),'
            f'IF(B{row}>0,1,0)))'
        )
        escribir_celda_segura(ws, row, 4, formula_cuentas)

    evolution_rows = max(1, report_end - report_start + 1)
    evolution_end = tablero_start + evolution_rows - 1

    cuentas = []
    for transaction in transactions:
        clave = (
            banco_corto(transaction["banco"]),
            str(transaction["cuenta"])
        )
        if clave not in cuentas:
            cuentas.append(clave)

    cuentas = sorted(cuentas, key=lambda x: (x[0], x[1]))

    cuenta_start = 15

    for row in range(cuenta_start, max(ws.max_row, cuenta_start + 100) + 1):
        for col in range(6, 9):
            limpiar_celda_segura(ws, row, col)

    cuenta_end = cuenta_start + len(cuentas) - 1
    total_row = cuenta_end + 1 if cuentas else cuenta_start
    required_account_end = max(cuenta_end, total_row)

    if required_account_end > ws.max_row:
        old_max = ws.max_row
        for row in range(old_max + 1, required_account_end + 1):
            copiar_estilo_fila(ws, 15, row)

    report_columns = {}
    banco_actual = ""

    for col in range(2, total_col):
        banco_cell = reporte.cell(row=3, column=col)
        cuenta_cell = reporte.cell(row=4, column=col)

        if not es_celda_combinada(banco_cell):
            banco_valor = banco_cell.value
            if banco_valor not in (None, ""):
                banco_actual = str(banco_valor).strip()

        if es_celda_combinada(cuenta_cell):
            continue

        cuenta = cuenta_cell.value
        if banco_actual and cuenta not in (None, ""):
            report_columns[
                (clean_text(banco_actual), str(cuenta).strip())
            ] = col

    for index, (banco, cuenta) in enumerate(cuentas, start=cuenta_start):
        escribir_celda_segura(ws, index, 6, valor_o_nd(banco))
        escribir_celda_segura(ws, index, 7, valor_o_nd(cuenta))

        report_col = report_columns.get((clean_text(banco), cuenta))

        if report_col:
            col_letter = get_column_letter(report_col)
            formula = (
                f'=IF(AND(OR($G$5="TODOS",TRIM($G$5)=TRIM(F{index})),'
                f'OR($E$5="TODAS",TEXT($E$5,"0")=TEXT(G{index},"0"))),'
                f'SUMIFS(\'{target_reporte}\'!${col_letter}$5:${col_letter}${report_end},'
                f'\'{target_reporte}\'!$A$5:$A${report_end},">="&$A$5,'
                f'\'{target_reporte}\'!$A$5:$A${report_end},"<="&$C$5),0)'
            )
            escribir_celda_segura(ws, index, 8, formula)
        else:
            escribir_celda_segura(ws, index, 8, 0)

    # Total visible debajo de CRÉDITOS POR CUENTA EN EL RANGO.
    escribir_celda_segura(ws, total_row, 6, "Total")
    escribir_celda_segura(ws, total_row, 7, None)

    if cuentas:
        escribir_celda_segura(
            ws,
            total_row,
            8,
            f'=SUM(H{cuenta_start}:H{cuenta_end})'
        )
    else:
        escribir_celda_segura(ws, total_row, 8, 0)

    # Estilo del Total basado en la tabla existente.
    try:
        from openpyxl.styles import PatternFill, Font

        for col in range(6, 9):
            source = ws.cell(row=cuenta_start, column=col)
            target = ws.cell(row=total_row, column=col)

            if not es_celda_combinada(source) and not es_celda_combinada(target):
                if source.has_style:
                    target._style = copy(source._style)
                target.fill = PatternFill(
                    fill_type="solid",
                    fgColor="E2F0D9"
                )
                target.font = Font(
                    name=source.font.name,
                    size=source.font.sz,
                    bold=True,
                    italic=source.font.italic,
                    color=source.font.color
                )
    except Exception:
        pass

    # KPIs superiores: se reescriben para evitar referencias heredadas
    # de otra hoja/moneda al copiar TABLERO CREDITOS.
    if cuentas:
        escribir_celda_segura(
            ws, 8, 1,
            f'=SUM(H{cuenta_start}:H{cuenta_end})'
        )
        escribir_celda_segura(
            ws, 8, 7,
            f'=COUNTIF(H{cuenta_start}:H{cuenta_end},">0")'
        )
    else:
        escribir_celda_segura(ws, 8, 1, 0)
        escribir_celda_segura(ws, 8, 7, 0)

    escribir_celda_segura(
        ws, 8, 3,
        f'=IFERROR(AVERAGEIF(B{tablero_start}:B{evolution_end},">0"),0)'
    )
    escribir_celda_segura(
        ws, 8, 5,
        f'=IFERROR(MAX(B{tablero_start}:B{evolution_end}),0)'
    )

    # SALDO DISPONIBLE en I:J.
    try:
        current_merges = {str(r) for r in ws.merged_cells.ranges}

        if "I7:J7" not in current_merges:
            ws.merge_cells("I7:J7")

        current_merges = {str(r) for r in ws.merged_cells.ranges}
        if "I8:J8" not in current_merges:
            ws.merge_cells("I8:J8")

        for source_coord, target_coord in (
            ("G7", "I7"),
            ("G8", "I8")
        ):
            source = ws[source_coord]
            target = ws[target_coord]

            if source.has_style:
                target._style = copy(source._style)
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.fill:
                target.fill = copy(source.fill)
            if source.font:
                target.font = copy(source.font)
            if source.border:
                target.border = copy(source.border)

        ws["I7"] = "SALDO DISPONIBLE"

    except Exception:
        escribir_celda_segura(ws, 7, 9, "SALDO DISPONIBLE")

    saldos_start = 5
    saldos_end = ultima_fila_con_valor(
        saldos,
        2,
        saldos_start
    )

    if saldos_end < saldos_start:
        saldo_formula = "=0"
    else:
        saldo_formula = (
            f'=IF($E$5<>"TODAS",'
            f'IFERROR(SUMIF(\'{target_saldos}\'!$B${saldos_start}:$B${saldos_end},'
            f'$E$5,\'{target_saldos}\'!$D${saldos_start}:$D${saldos_end}),0),'
            f'IF($G$5<>"TODOS",'
            f'IFERROR(SUMIF(\'{target_saldos}\'!$A${saldos_start}:$A${saldos_end},'
            f'$G$5,\'{target_saldos}\'!$D${saldos_start}:$D${saldos_end}),0),'
            f'IFERROR(SUM(\'{target_saldos}\'!$D${saldos_start}:$D${saldos_end}),0)))'
        )

    escribir_celda_segura(ws, 8, 9, saldo_formula)

    actualizar_graficas_dinamicas(
        ws,
        evolution_end,
        cuenta_end if cuentas else cuenta_start,
        len(cuentas)
    )


# ============================================================
# PROCESAR PLANTILLA COMPLETA
# ============================================================



def aplicar_formato_juego_moneda(
    workbook,
    transactions,
    moneda,
    nombres
):
    """
    Aplica una sola moneda a TODO el juego de hojas ya separado:
    saldos, estados, reporte diario y tablero.
    No convierte importes; solo corrige el formato visual.
    """
    formato = formato_moneda_excel(moneda)

    # ESTADOS CONSOLIDADOS
    ws_estados = workbook[nombres[SHEET_ESTADOS]]
    for row_number in range(2, 2 + len(transactions)):
        for column in (6, 7, 8):
            cell = obtener_celda_segura(ws_estados, row_number, column)
            if cell:
                cell.number_format = formato

    # SALDOS POR CUENTA
    ws_saldos = workbook[nombres[SHEET_SALDOS]]
    cuentas = []
    for t in transactions:
        clave = (str(t.get("banco", "")).strip(), str(t.get("cuenta", "")).strip())
        if clave not in cuentas:
            cuentas.append(clave)

    for row_number in range(5, 5 + len(cuentas)):
        for column in (3, 4):
            cell = obtener_celda_segura(ws_saldos, row_number, column)
            if cell:
                cell.number_format = formato

    # REPORTE CRÉDITOS DIARIOS
    ws_reporte = workbook[nombres[SHEET_REPORTE]]
    fechas = {
        (
            t["fecha"].date()
            if isinstance(t.get("fecha"), datetime)
            else t.get("fecha")
        )
        for t in transactions
        if isinstance(t.get("fecha"), (datetime, date))
    }
    last_data_row = 4 + len(fechas) if fechas else 4

    # Detectar TOTAL y CUENTAS CON ABONO.
    total_col = None
    count_col = None
    for col in range(2, ws_reporte.max_column + 1):
        header = clean_text(ws_reporte.cell(row=4, column=col).value)
        if total_col is None and "total" in header and ("credito" in header or "credit" in header):
            total_col = col
        if count_col is None and "cuentas" in header and "abono" in header:
            count_col = col

    # Todas las columnas de cuentas y el total son de la misma moneda.
    last_money_col = total_col if total_col is not None else max(2, ws_reporte.max_column - 1)
    for row in range(5, last_data_row + 1):
        for col in range(2, last_money_col + 1):
            cell = obtener_celda_segura(ws_reporte, row, col)
            if cell:
                cell.number_format = formato

        if count_col is not None:
            cell = obtener_celda_segura(ws_reporte, row, count_col)
            if cell:
                cell.number_format = "0"

    # TABLERO CRÉDITOS
    ws_tablero = workbook[nombres[SHEET_TABLERO]]

    # KPIs superiores: Créditos en rango, Promedio diario y Mayor día.
    titulos_monetarios = (
        "creditos en rango",
        "promedio diario",
        "mayor dia",
        "saldo disponible",
    )
    titulo_cantidad = "cuentas con abono"

    for row in range(1, min(ws_tablero.max_row, 14) + 1):
        for column in range(1, ws_tablero.max_column + 1):
            cell = ws_tablero.cell(row=row, column=column)
            if es_celda_combinada(cell):
                continue

            texto = clean_text(cell.value)
            if not texto:
                continue

            if any(titulo in texto for titulo in titulos_monetarios):
                value_cell = obtener_celda_segura(ws_tablero, row + 1, column)
                if value_cell:
                    value_cell.number_format = formato
            elif titulo_cantidad in texto:
                value_cell = obtener_celda_segura(ws_tablero, row + 1, column)
                if value_cell:
                    value_cell.number_format = "0"

    # Evolución diaria: crédito seleccionado y variación diaria.
    for row in range(15, ws_tablero.max_row + 1):
        for column in (2, 3):
            cell = obtener_celda_segura(ws_tablero, row, column)
            if cell:
                cell.number_format = formato

    # Créditos por cuenta: columna H.
    for row in range(15, 15 + len(cuentas)):
        cell = obtener_celda_segura(ws_tablero, row, 8)
        if cell:
            cell.number_format = formato

    # Total de créditos por cuenta y Saldo Disponible.
    total_row = 15 + len(cuentas)

    total_cell = obtener_celda_segura(ws_tablero, total_row, 8)
    if total_cell:
        total_cell.number_format = formato

    saldo_disponible_cell = obtener_celda_segura(ws_tablero, 8, 9)
    if saldo_disponible_cell:
        saldo_disponible_cell.number_format = formato

    # Eje monetario de las gráficas.
    for chart in getattr(ws_tablero, "_charts", []):
        try:
            if hasattr(chart, "y_axis"):
                chart.y_axis.numFmt = formato
        except Exception:
            pass


def crear_versiones_por_moneda(workbook, transactions):
    """
    Separa físicamente los reportes por moneda.

    - Las hojas base conservan la moneda principal (GTQ si existe).
    - Para cada moneda adicional se COPIAN las hojas base para conservar
      formato, anchos, altos, celdas combinadas y configuración visual.
    - Cada juego de hojas se llena únicamente con las transacciones de
      su propia moneda.
    """
    monedas = sorted({
        normalizar_codigo_moneda(t.get("moneda"))
        for t in transactions
        if normalizar_codigo_moneda(t.get("moneda")) not in {"N/D", ""}
    })

    if len(monedas) <= 1:
        return

    moneda_base = "GTQ" if "GTQ" in monedas else monedas[0]
    monedas_adicionales = [m for m in monedas if m != moneda_base]

    # Crear TODAS las copias antes de modificar las hojas base.
    # Así cada moneda parte de la plantilla original, no de datos ya escritos.
    nombres_por_moneda = {
        moneda_base: {
            SHEET_SALDOS: SHEET_SALDOS,
            SHEET_ESTADOS: SHEET_ESTADOS,
            SHEET_REPORTE: SHEET_REPORTE,
            SHEET_TABLERO: SHEET_TABLERO,
        }
    }

    for moneda in monedas_adicionales:
        nombres = {}
        for base_name in (
            SHEET_SALDOS,
            SHEET_ESTADOS,
            SHEET_REPORTE,
            SHEET_TABLERO
        ):
            nuevo_nombre = f"{base_name} {moneda}"

            if nuevo_nombre in workbook.sheetnames:
                del workbook[nuevo_nombre]

            hoja_copia = workbook.copy_worksheet(
                workbook[base_name]
            )
            hoja_copia.title = nuevo_nombre
            nombres[base_name] = nuevo_nombre

        nombres_por_moneda[moneda] = nombres

    # Escribir cada moneda exclusivamente en su propio juego de hojas.
    for moneda in [moneda_base] + monedas_adicionales:
        trans_moneda = [
            t for t in transactions
            if normalizar_codigo_moneda(t.get("moneda")) == moneda
        ]

        if not trans_moneda:
            continue

        nombres = nombres_por_moneda[moneda]

        escribir_saldos_por_cuenta(
            workbook,
            trans_moneda,
            sheet_name=nombres[SHEET_SALDOS]
        )

        escribir_estados_consolidados(
            workbook,
            trans_moneda,
            sheet_name=nombres[SHEET_ESTADOS]
        )

        escribir_reporte_creditos(
            workbook,
            trans_moneda,
            sheet_name=nombres[SHEET_REPORTE]
        )

        actualizar_tablero(
            workbook,
            trans_moneda,
            sheet_name=nombres[SHEET_TABLERO],
            reporte_sheet_name=nombres[SHEET_REPORTE],
            saldos_sheet_name=nombres[SHEET_SALDOS]
        )

        # Aplicar la moneda correcta a TODO este juego de hojas.
        aplicar_formato_juego_moneda(
            workbook,
            trans_moneda,
            moneda,
            nombres
        )


def generar_por_moneda_si_aplica(workbook, transactions):
    monedas = {
        normalizar_codigo_moneda(t.get("moneda"))
        for t in transactions
        if normalizar_codigo_moneda(t.get("moneda")) not in {"N/D", ""}
    }

    return len(monedas) > 1


def insertar_en_plantilla(
    transactions,
    output_path
):

    if not os.path.exists(
        TEMPLATE_FILE
    ):

        raise Exception(
            "No existe la plantilla: "
            + TEMPLATE_FILE
        )

    workbook = load_workbook(
        TEMPLATE_FILE
    )

    try:

        # Si hay varias monedas, NO generar primero las hojas mixtas.
        # Cada moneda tendrá su propio juego de reportes.
        monedas = {
            normalizar_codigo_moneda(t.get("moneda"))
            for t in transactions
            if normalizar_codigo_moneda(t.get("moneda")) not in {"N/D", ""}
        }

        if len(monedas) > 1:
            crear_versiones_por_moneda(workbook, transactions)
        else:
            escribir_estados_consolidados(workbook, transactions)
            escribir_saldos_por_cuenta(workbook, transactions)
            escribir_reporte_creditos(workbook, transactions)
            actualizar_tablero(workbook, transactions)

        # 5
        # Aplicar formato global únicamente cuando el libro tiene UNA moneda.
        # Si hay varias monedas, crear_versiones_por_moneda ya aplicó el
        # formato correcto de forma independiente a cada juego de hojas.
        if len(monedas) <= 1:
            aplicar_formato_moneda_excel(
                workbook,
                transactions
            )

        # Recalcular fórmulas/gráficas al abrir el resultado.
        forzar_recalculo_excel(workbook)

        workbook.save(
            output_path
        )

    finally:

        workbook.close()


# ============================================================
# DATOS PARA DASHBOARD HTML
# ============================================================

def fecha_dashboard(value):
    """Fecha legible dd/mm/YYYY."""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    return str(value or "")


def fecha_dashboard_iso(value):
    """Fecha ISO YYYY-mm-dd utilizada por los filtros del HTML."""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    texto = str(value or "").strip()

    if not texto:
        return ""

    for formato in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y"
    ):
        try:
            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except Exception:
            pass

    return ""


def _redondear_dashboard(value):
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _clave_banco_dashboard(banco):
    """
    Normaliza las llaves que utiliza la plantilla para colores/logos.
    Para bancos no conocidos conserva una llave estable en mayúsculas.
    """

    original = str(banco or "BANCO NO IDENTIFICADO").strip()
    texto = clean_text(original)

    if "agricola" in texto:
        return "BANCO AGRÍCOLA"

    if "cuscatlan" in texto:
        return "BANCO CUSCATLÁN"

    if re.search(r"\bbac\b", texto) or "america central" in texto:
        return "BAC"

    if "promerica" in texto:
        return "PROMERICA"

    if "davivienda" in texto:
        return "DAVIVIENDA"

    if "industrial" in texto:
        return "BANCO INDUSTRIAL"

    if "banrural" in texto or "desarrollo rural" in texto:
        return "BANRURAL"

    if "g&t" in original.lower() or "gyt" in texto:
        return "G&T"

    if texto == "bam" or "agricola mercantil" in texto:
        return "BAM"

    if "ficohsa" in texto:
        return "FICOHSA"

    return original.upper() or "BANCO NO IDENTIFICADO"


def _nombre_banco_dashboard(clave, original=""):
    nombres = {
        "BANCO AGRÍCOLA": "Banco Agrícola",
        "BANCO CUSCATLÁN": "Banco Cuscatlán",
        "BAC": "BAC",
        "PROMERICA": "Promerica",
        "DAVIVIENDA": "Davivienda",
        "BANCO INDUSTRIAL": "Banco Industrial",
        "BANRURAL": "Banrural",
        "G&T": "G&T Continental",
        "BAM": "BAM",
        "FICOHSA": "Ficohsa",
    }

    if clave in nombres:
        return nombres[clave]

    texto = str(original or clave or "Banco").strip()
    return texto if texto else "Banco"


def _color_banco_dashboard(clave, indice=0):
    colores_conocidos = {
        "BANCO AGRÍCOLA": "#18B981",
        "BANCO CUSCATLÁN": "#8B5CF6",
        "BAC": "#F05252",
        "PROMERICA": "#3B82F6",
        "DAVIVIENDA": "#F59E0B",
        "BANCO INDUSTRIAL": "#2563EB",
        "BANRURAL": "#16A34A",
        "G&T": "#7C3AED",
        "BAM": "#0891B2",
        "FICOHSA": "#DC2626",
    }

    if clave in colores_conocidos:
        return colores_conocidos[clave]

    paleta = [
        "#60A5FA",
        "#34D399",
        "#A78BFA",
        "#FBBF24",
        "#FB7185",
        "#22D3EE",
        "#F97316",
        "#84CC16",
        "#E879F9",
        "#94A3B8",
    ]

    return paleta[indice % len(paleta)]


def _cuenta_corta_dashboard(cuenta):
    texto = str(cuenta or "N/D").strip()
    digitos = re.sub(r"[^0-9]", "", texto)

    if len(digitos) > 4:
        return digitos[-4:]

    return digitos or texto or "N/D"


def _fecha_corte_desde_nombre(nombre_archivo, fecha_respaldo=""):
    """Intenta obtener una fecha de corte del nombre; si no, usa la última fecha."""

    nombre = str(nombre_archivo or "")

    patrones = [
        (r"(?<!\d)(\d{2})(\d{2})(20\d{2})(?!\d)", "%d/%m/%Y"),
        (r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)", "%Y/%m/%d"),
        (r"(?<!\d)(\d{2})[-_.](\d{2})[-_.](20\d{2})(?!\d)", "%d/%m/%Y"),
    ]

    for patron, tipo in patrones:
        match = re.search(patron, nombre)
        if not match:
            continue

        try:
            if tipo == "%Y/%m/%d":
                year, month, day = match.groups()
            else:
                day, month, year = match.groups()

            return datetime(
                int(year),
                int(month),
                int(day)
            ).strftime("%d/%m/%Y")

        except Exception:
            pass

    return fecha_respaldo or ""



def resolver_moneda_dashboard_final(
    transactions,
    nombre_archivo="",
    moneda_contextual="N/D"
):
    """
    Resuelve UNA moneda de presentación para el dashboard.

    Prioridad:
    1. Moneda contextual/textual confiable del archivo.
    2. País indicado en el nombre del archivo.
    3. Conjunto de bancos del archivo.
    4. Moneda resultante de las transacciones.

    Esto corrige falsos MULTI causados por formatos heredados de Excel.
    No convierte importes.
    """

    contextual = normalizar_codigo_moneda(
        moneda_contextual
    )

    if contextual not in {"N/D", ""}:
        return contextual

    archivo = clean_text(
        nombre_archivo
    )

    if "el salvador" in archivo:
        return "USD"

    if "guatemala" in archivo:
        return "GTQ"

    if "honduras" in archivo:
        return "HNL"

    if "costa rica" in archivo:
        return "CRC"

    if "nicaragua" in archivo:
        return "NIO"

    if "mexico" in archivo:
        return "MXN"

    bancos_texto = " | ".join(
        str(
            transaction.get(
                "banco",
                ""
            )
        )
        for transaction in transactions
    )

    bancos_normal = clean_text(
        bancos_texto
    )

    # Señales muy fuertes de Guatemala.
    guatemala_signals = (
        "banco industrial",
        "industrial",
        "banrural",
        "g&t",
        "gyt",
        "banco agricola mercantil",
        "bam",
    )

    if any(
        signal in bancos_normal
        for signal in guatemala_signals
    ):
        return "GTQ"

    # Conjunto típico de El Salvador.
    tiene_agricola = (
        "banco agricola" in bancos_normal
        or "agricola" in bancos_normal
    )

    tiene_cuscatlan = (
        "cuscatlan" in bancos_normal
    )

    if tiene_agricola and tiene_cuscatlan:
        return "USD"

    detectada = moneda_unica_transacciones(
        transactions
    )

    if detectada in {
        "GTQ", "USD", "EUR", "SVC",
        "HNL", "CRC", "NIO", "MXN", "GBP"
    }:
        return detectada

    return "N/D"



def construir_dashboard_data(
    transactions,
    nombre_archivo="",
    moneda_preferida="N/D"
):
    """
    Construye el dashboard con separación real por país.

    Cada banco/cuenta conserva:
      - country: GUATEMALA / EL_SALVADOR / NO_IDENTIFICADO
      - currency: moneda detectada en su hoja
      - bankKey: llave única PAIS|BANCO

    De esta forma BAC, Promerica u otros bancos con nombres regionales
    nunca se mezclan entre Guatemala y El Salvador.

    No se convierten monedas.
    """

    transactions = list(
        transactions or []
    )

    if not transactions:
        data = json.loads(
            json.dumps(
                DASHBOARD_CACHE
            )
        )
        data["meta"]["file"] = (
            nombre_archivo or ""
        )
        data["meta"]["countries"] = []
        return data

    # Asegurar país incluso para transacciones provenientes de una versión
    # anterior del parser.
    for transaction in transactions:
        pais = normalizar_pais_bancario(
            transaction.get(
                "pais"
            )
        )

        if pais == "NO_IDENTIFICADO":
            pais = detectar_pais_bancario(
                banco=transaction.get(
                    "banco",
                    ""
                ),
                moneda=transaction.get(
                    "moneda",
                    "N/D"
                ),
                nombre_archivo=nombre_archivo
            )

        transaction["pais"] = pais

    # --------------------------------------------------------
    # MONEDA GENERAL DEL ARCHIVO
    # --------------------------------------------------------
    paises_presentes = {
        normalizar_pais_bancario(
            transaction.get(
                "pais"
            )
        )
        for transaction in transactions
    }

    paises_presentes.discard(
        "NO_IDENTIFICADO"
    )

    if len(paises_presentes) > 1:
        # Nunca usar una sola moneda para sumar/mostrar dos países.
        moneda_dashboard = "MULTI"
    else:
        moneda_dashboard = resolver_moneda_dashboard_final(
            transactions,
            nombre_archivo=nombre_archivo,
            moneda_contextual=moneda_preferida
        )

    # --------------------------------------------------------
    # IDENTIFICAR CUENTAS SIN COLISIONES ENTRE PAÍSES
    # --------------------------------------------------------
    originales_por_banco = {}
    orden_cuentas = []
    banco_meta = {}

    for transaction in transactions:
        banco_original = str(
            transaction.get(
                "banco",
                "BANCO NO IDENTIFICADO"
            )
        ).strip()

        canonical = _clave_banco_dashboard(
            banco_original
        )

        pais = normalizar_pais_bancario(
            transaction.get(
                "pais"
            )
        )

        bank_key = _clave_banco_pais_dashboard(
            pais,
            canonical
        )

        cuenta_original = str(
            transaction.get(
                "cuenta",
                "N/D"
            )
        ).strip()

        clave = (
            bank_key,
            cuenta_original
        )

        if clave not in orden_cuentas:
            orden_cuentas.append(
                clave
            )

        originales_por_banco.setdefault(
            bank_key,
            []
        )

        if (
            cuenta_original
            not in originales_por_banco[
                bank_key
            ]
        ):
            originales_por_banco[
                bank_key
            ].append(
                cuenta_original
            )

        banco_meta.setdefault(
            bank_key,
            {
                "canonical": canonical,
                "country": pais,
                "original": banco_original
            }
        )

    cuenta_display = {}

    for (
        bank_key,
        cuentas_originales
    ) in originales_por_banco.items():

        candidatos = {}

        for cuenta_original in cuentas_originales:
            corta = _cuenta_corta_dashboard(
                cuenta_original
            )

            candidatos.setdefault(
                corta,
                []
            ).append(
                cuenta_original
            )

        for (
            corta,
            originales
        ) in candidatos.items():

            if len(originales) == 1:
                cuenta_display[
                    (
                        bank_key,
                        originales[0]
                    )
                ] = corta
                continue

            usados = set()

            for cuenta_original in originales:
                digitos = re.sub(
                    r"[^0-9]",
                    "",
                    cuenta_original
                )

                visible = (
                    digitos[-6:]
                    if len(digitos) > 6
                    else (
                        digitos
                        or cuenta_original
                    )
                )

                if visible in usados:
                    visible = (
                        digitos
                        or cuenta_original
                    )

                usados.add(
                    visible
                )

                cuenta_display[
                    (
                        bank_key,
                        cuenta_original
                    )
                ] = visible

    # --------------------------------------------------------
    # TRANSACCIONES CRUDAS DEL DASHBOARD
    # --------------------------------------------------------
    raw_transactions = []
    grupos_cuenta = {}
    fechas_validas = []

    for (
        indice,
        transaction
    ) in enumerate(
        transactions,
        start=2
    ):
        banco_original = str(
            transaction.get(
                "banco",
                "BANCO NO IDENTIFICADO"
            )
        ).strip()

        canonical = _clave_banco_dashboard(
            banco_original
        )

        pais = normalizar_pais_bancario(
            transaction.get(
                "pais"
            )
        )

        bank_key = _clave_banco_pais_dashboard(
            pais,
            canonical
        )

        banco_nombre = _nombre_banco_dashboard(
            canonical,
            banco_original
        )

        cuenta_original = str(
            transaction.get(
                "cuenta",
                "N/D"
            )
        ).strip()

        cuenta_visible = cuenta_display.get(
            (
                bank_key,
                cuenta_original
            ),
            _cuenta_corta_dashboard(
                cuenta_original
            )
        )

        fecha_iso = fecha_dashboard_iso(
            transaction.get(
                "fecha"
            )
        )

        if fecha_iso:
            fechas_validas.append(
                fecha_iso
            )

        debito = abs(
            float(
                transaction.get(
                    "debito",
                    0
                )
                or 0
            )
        )

        credito = abs(
            float(
                transaction.get(
                    "credito",
                    0
                )
                or 0
            )
        )

        saldo = None

        if transaction.get(
            "_tiene_saldo"
        ):
            saldo = _redondear_dashboard(
                transaction.get(
                    "saldo",
                    0
                )
            )

        moneda = normalizar_codigo_moneda(
            transaction.get(
                "moneda",
                "N/D"
            )
        )

        fila = {
            "row": indice,
            "bankKey": bank_key,
            "bankCanonical": canonical,
            "bank": banco_nombre,
            "country": pais,
            "currency": moneda,
            "account": cuenta_visible,
            "original": cuenta_original,
            "date": fecha_iso,
            "debit": _redondear_dashboard(
                debito
            ),
            "credit": _redondear_dashboard(
                credito
            ),
            "balance": saldo,
            "reference": str(
                transaction.get(
                    "referencia",
                    ""
                )
                or ""
            ),
            "description": str(
                transaction.get(
                    "descripcion",
                    ""
                )
                or ""
            )
        }

        raw_transactions.append(
            fila
        )

        clave_cuenta = (
            bank_key,
            cuenta_original
        )

        grupos_cuenta.setdefault(
            clave_cuenta,
            []
        ).append({
            "order": indice,
            "date": fecha_iso,
            "debit": debito,
            "credit": credito,
            "balance": saldo,
            "currency": moneda,
            "country": pais,
            "saldo_inicial_cuenta": transaction.get(
                "saldo_inicial_cuenta"
            ),
            "has_balance": bool(
                transaction.get(
                    "_tiene_saldo"
                )
            )
        })

    raw_transactions = [
        item
        for item in raw_transactions
        if item.get(
            "date"
        )
    ]

    # --------------------------------------------------------
    # CUENTAS
    # --------------------------------------------------------
    accounts = []

    for (
        bank_key,
        cuenta_original
    ) in orden_cuentas:

        movimientos = grupos_cuenta.get(
            (
                bank_key,
                cuenta_original
            ),
            []
        )

        movimientos = sorted(
            movimientos,
            key=lambda item: (
                item.get(
                    "date",
                    ""
                ),
                item.get(
                    "order",
                    0
                )
            )
        )

        if not movimientos:
            continue

        primero = movimientos[0]

        saldo_inicial_explicito = primero.get(
            "saldo_inicial_cuenta"
        )

        if saldo_inicial_explicito is not None:
            initial = _redondear_dashboard(
                saldo_inicial_explicito
            )

        elif (
            primero.get(
                "has_balance"
            )
            and primero.get(
                "balance"
            )
            is not None
        ):
            initial = _redondear_dashboard(
                float(
                    primero.get(
                        "balance"
                    )
                    or 0
                )
                + float(
                    primero.get(
                        "debit"
                    )
                    or 0
                )
                - float(
                    primero.get(
                        "credit"
                    )
                    or 0
                )
            )

        else:
            initial = 0.0

        saldo_final = None

        for movimiento in movimientos:
            if (
                movimiento.get(
                    "has_balance"
                )
                and movimiento.get(
                    "balance"
                )
                is not None
            ):
                saldo_final = movimiento.get(
                    "balance"
                )

        creditos = sum(
            float(
                item.get(
                    "credit"
                )
                or 0
            )
            for item in movimientos
        )

        debitos = sum(
            float(
                item.get(
                    "debit"
                )
                or 0
            )
            for item in movimientos
        )

        if saldo_final is None:
            saldo_final = (
                initial
                + creditos
                - debitos
            )

        final = _redondear_dashboard(
            saldo_final
        )

        change = _redondear_dashboard(
            final
            - initial
        )

        change_pct = (
            _redondear_dashboard(
                change
                / initial
                * 100
            )
            if initial != 0
            else None
        )

        meta_banco = banco_meta.get(
            bank_key,
            {}
        )

        canonical = meta_banco.get(
            "canonical",
            bank_key.split(
                "|",
                1
            )[-1]
        )

        pais = meta_banco.get(
            "country",
            primero.get(
                "country",
                "NO_IDENTIFICADO"
            )
        )

        banco_original = meta_banco.get(
            "original",
            canonical
        )

        moneda_cuenta = moneda_pais_dashboard(
            pais,
            [
                item.get(
                    "currency"
                )
                for item in movimientos
            ]
        )

        accounts.append({
            "bankKey": bank_key,
            "bankCanonical": canonical,
            "bank": _nombre_banco_dashboard(
                canonical,
                banco_original
            ),
            "country": pais,
            "currency": moneda_cuenta,
            "account": cuenta_display.get(
                (
                    bank_key,
                    cuenta_original
                ),
                _cuenta_corta_dashboard(
                    cuenta_original
                )
            ),
            "original": cuenta_original,
            "initial": initial,
            "final": final,
            "change": change,
            "changePct": change_pct,
            "credits": _redondear_dashboard(
                creditos
            ),
            "debits": _redondear_dashboard(
                debitos
            ),
            "movements": len(
                movimientos
            )
        })

    # --------------------------------------------------------
    # BANCOS
    # --------------------------------------------------------
    bank_keys = []

    for account in accounts:
        if (
            account[
                "bankKey"
            ]
            not in bank_keys
        ):
            bank_keys.append(
                account[
                    "bankKey"
                ]
            )

    total_final_por_pais = {}

    for account in accounts:
        pais = account.get(
            "country",
            "NO_IDENTIFICADO"
        )

        total_final_por_pais[
            pais
        ] = (
            total_final_por_pais.get(
                pais,
                0.0
            )
            + float(
                account.get(
                    "final"
                )
                or 0
            )
        )

    banks = []

    for (
        indice,
        bank_key
    ) in enumerate(
        bank_keys
    ):

        cuentas_banco = [
            account
            for account in accounts
            if account[
                "bankKey"
            ]
            == bank_key
        ]

        trans_banco = [
            item
            for item in raw_transactions
            if item[
                "bankKey"
            ]
            == bank_key
        ]

        if not cuentas_banco:
            continue

        pais = cuentas_banco[
            0
        ].get(
            "country",
            "NO_IDENTIFICADO"
        )

        canonical = cuentas_banco[
            0
        ].get(
            "bankCanonical",
            bank_key.split(
                "|",
                1
            )[-1]
        )

        initial = _redondear_dashboard(
            sum(
                account[
                    "initial"
                ]
                for account
                in cuentas_banco
            )
        )

        final = _redondear_dashboard(
            sum(
                account[
                    "final"
                ]
                for account
                in cuentas_banco
            )
        )

        credits = _redondear_dashboard(
            sum(
                item[
                    "credit"
                ]
                for item
                in trans_banco
            )
        )

        debits = _redondear_dashboard(
            sum(
                item[
                    "debit"
                ]
                for item
                in trans_banco
            )
        )

        change = _redondear_dashboard(
            final
            - initial
        )

        change_pct = (
            _redondear_dashboard(
                change
                / initial
                * 100
            )
            if initial != 0
            else None
        )

        net_flow = _redondear_dashboard(
            credits
            - debits
        )

        total_pais = _redondear_dashboard(
            total_final_por_pais.get(
                pais,
                0.0
            )
        )

        share = (
            _redondear_dashboard(
                final
                / total_pais
                * 100
            )
            if total_pais != 0
            else 0.0
        )

        meta_banco = banco_meta.get(
            bank_key,
            {}
        )

        banco_original = meta_banco.get(
            "original",
            canonical
        )

        moneda_banco = moneda_pais_dashboard(
            pais,
            [
                item.get(
                    "currency"
                )
                for item
                in trans_banco
            ]
        )

        banks.append({
            "key": bank_key,
            "canonicalKey": canonical,
            "name": _nombre_banco_dashboard(
                canonical,
                banco_original
            ),
            "country": pais,
            "currency": moneda_banco,
            "color": _color_banco_dashboard(
                canonical,
                indice
            ),
            "initial": initial,
            "final": final,
            "change": change,
            "changePct": change_pct,
            "credits": credits,
            "debits": debits,
            "netFlow": net_flow,
            "share": share,
            "accounts": len(
                cuentas_banco
            ),
            "movements": len(
                trans_banco
            )
        })

    # --------------------------------------------------------
    # RESUMEN DIARIO GLOBAL
    # El HTML lo vuelve a calcular por país y por filtro de fecha.
    # --------------------------------------------------------
    por_fecha = {}

    for item in raw_transactions:
        fecha_iso = item.get(
            "date"
        )

        if not fecha_iso:
            continue

        registro = por_fecha.setdefault(
            fecha_iso,
            {
                "date": fecha_iso,
                "moves": 0,
                "debits": 0.0,
                "credits": 0.0
            }
        )

        registro[
            "moves"
        ] += 1

        registro[
            "debits"
        ] += float(
            item.get(
                "debit"
            )
            or 0
        )

        registro[
            "credits"
        ] += float(
            item.get(
                "credit"
            )
            or 0
        )

    daily = []

    for fecha_iso in sorted(
        por_fecha.keys()
    ):
        item = por_fecha[
            fecha_iso
        ]

        debits = _redondear_dashboard(
            item[
                "debits"
            ]
        )

        credits = _redondear_dashboard(
            item[
                "credits"
            ]
        )

        try:
            label = datetime.strptime(
                fecha_iso,
                "%Y-%m-%d"
            ).strftime(
                "%d/%m"
            )

        except Exception:
            label = fecha_iso

        daily.append({
            "date": fecha_iso,
            "label": label,
            "moves": int(
                item[
                    "moves"
                ]
            ),
            "debits": debits,
            "credits": credits,
            "net": _redondear_dashboard(
                credits
                - debits
            )
        })

    max_credit = (
        max(
            daily,
            key=lambda item: item[
                "credits"
            ]
        )
        if daily
        else None
    )

    max_debit = (
        max(
            daily,
            key=lambda item: item[
                "debits"
            ]
        )
        if daily
        else None
    )

    max_net = (
        max(
            daily,
            key=lambda item: item[
                "net"
            ]
        )
        if daily
        else None
    )

    min_net = (
        min(
            daily,
            key=lambda item: item[
                "net"
            ]
        )
        if daily
        else None
    )

    # --------------------------------------------------------
    # RESÚMENES POR PAÍS
    # --------------------------------------------------------
    country_summaries = []

    for pais in (
        "GUATEMALA",
        "EL_SALVADOR",
        "NO_IDENTIFICADO"
    ):
        trans_pais = [
            item
            for item in raw_transactions
            if item.get(
                "country"
            )
            == pais
        ]

        cuentas_pais = [
            item
            for item in accounts
            if item.get(
                "country"
            )
            == pais
        ]

        bancos_pais = [
            item
            for item in banks
            if item.get(
                "country"
            )
            == pais
        ]

        if not (
            trans_pais
            or cuentas_pais
            or bancos_pais
        ):
            continue

        currency = moneda_pais_dashboard(
            pais,
            [
                item.get(
                    "currency"
                )
                for item in trans_pais
            ]
        )

        country_summaries.append({
            "key": pais,
            "label": (
                "Guatemala"
                if pais == "GUATEMALA"
                else (
                    "El Salvador"
                    if pais == "EL_SALVADOR"
                    else "Sin clasificar"
                )
            ),
            "currency": currency,
            "banks": len(
                bancos_pais
            ),
            "accounts": len(
                cuentas_pais
            ),
            "movementCount": len(
                trans_pais
            )
        })

    # --------------------------------------------------------
    # META GENERAL
    # --------------------------------------------------------
    total_initial = _redondear_dashboard(
        sum(
            account[
                "initial"
            ]
            for account
            in accounts
        )
    )

    total_final = _redondear_dashboard(
        sum(
            account[
                "final"
            ]
            for account
            in accounts
        )
    )

    total_credits = _redondear_dashboard(
        sum(
            item[
                "credit"
            ]
            for item
            in raw_transactions
        )
    )

    total_debits = _redondear_dashboard(
        sum(
            item[
                "debit"
            ]
            for item
            in raw_transactions
        )
    )

    net_flow = _redondear_dashboard(
        total_credits
        - total_debits
    )

    change_pct = (
        _redondear_dashboard(
            net_flow
            / total_debits
            * 100
        )
        if total_debits != 0
        else 0.0
    )

    latest_iso = (
        max(
            fechas_validas
        )
        if fechas_validas
        else ""
    )

    latest_display = ""

    if latest_iso:
        try:
            latest_display = datetime.strptime(
                latest_iso,
                "%Y-%m-%d"
            ).strftime(
                "%d/%m/%Y"
            )

        except Exception:
            latest_display = latest_iso

    file_cut = _fecha_corte_desde_nombre(
        nombre_archivo,
        latest_display
    )

    return {
        "meta": {
            "file": nombre_archivo or "Archivo bancario",
            "fileCut": file_cut,
            "latestMovement": latest_display,
            "currency": moneda_dashboard,
            "accounts": len(
                accounts
            ),
            "banks": len(
                banks
            ),
            "movementCount": len(
                raw_transactions
            ),
            "mixedCountries": (
                len(
                    paises_presentes
                )
                > 1
            ),
            "countries": country_summaries,
            "generatedAt": datetime.now().strftime(
                "%d/%m/%Y %H:%M"
            )
        },
        "totals": {
            "initial": total_initial,
            "final": total_final,
            "change": net_flow,
            "changePct": change_pct,
            "credits": total_credits,
            "debits": total_debits,
            "netFlow": net_flow
        },
        "highlights": {
            "maxCredit": max_credit,
            "maxDebit": max_debit,
            "maxNet": max_net,
            "minNet": min_net
        },
        "banks": banks,
        "accounts": accounts,
        "daily": daily,
        "transactions": raw_transactions
    }



def guardar_dashboard_data(
    data
):

    global DASHBOARD_CACHE

    DASHBOARD_CACHE = data

    try:

        temporal = (
            DASHBOARD_DATA_FILE
            + ".tmp"
        )

        with open(
            temporal,
            "w",
            encoding="utf-8"
        ) as archivo:

            json.dump(
                data,
                archivo,
                ensure_ascii=False,
                indent=2
            )

        os.replace(
            temporal,
            DASHBOARD_DATA_FILE
        )

    except Exception as error:

        print(
            "No se pudo guardar dashboard_data.json:",
            str(error)
        )


def cargar_dashboard_data():

    try:

        if os.path.exists(
            DASHBOARD_DATA_FILE
        ):

            with open(
                DASHBOARD_DATA_FILE,
                "r",
                encoding="utf-8"
            ) as archivo:

                return json.load(
                    archivo
                )

    except Exception as error:

        print(
            "Error leyendo dashboard_data.json:",
            str(error)
        )

    return DASHBOARD_CACHE



def _etiqueta_moneda_dashboard(codigo):
    codigo = str(codigo or "N/D").upper()
    etiquetas = {
        "USD": "USD ($)",
        "GTQ": "GTQ (Q)",
        "EUR": "EUR (€)",
        "SVC": "SVC (₡)",
        "HNL": "HNL (L)",
        "CRC": "CRC (₡)",
        "NIO": "NIO (C$)",
        "MXN": "MXN (MX$)",
        "GBP": "GBP (£)",
        "MULTI": "Múltiples monedas",
        "N/D": "N/D"
    }
    return etiquetas.get(codigo, codigo)


def renderizar_dashboard_nueva_plantilla(datos_override=None):
    """
    Renderiza templates/dashboard/index.html y sustituye automáticamente
    los dos bloques de datos fijos que trae la plantilla nueva:

        const BASE_DATA = {...};
        const RAW_TRANSACTIONS = [...];

    De esta forma la plantilla puede conservar todo su diseño, filtros,
    pestañas, gráficos y logos, pero siempre usa el último Excel procesado.
    """

    if not os.path.exists(DASHBOARD_TEMPLATE):
        raise FileNotFoundError(
            "No existe index.html del dashboard: "
            + DASHBOARD_TEMPLATE
        )

    datos = (
        datos_override
        if datos_override is not None
        else cargar_dashboard_data()
    )

    # Jinja inyecta los datos del dashboard para que la plantilla
    # quede limpia y la lógica del cliente viva en assets/js/dashboard.js.
    html = render_template(
        "dashboard/index.html",
        dashboard_data=datos,
        base_data=datos,
        raw_transactions=datos.get("transactions", [])
    )

    return html



def _dashboard_asset_path(url):
    """
    Convierte una URL /static/... del HTML en la ruta física correspondiente
    dentro del proyecto Flask. Solo permite archivos dentro de BASE_DIR/static.
    """
    url = str(url or "").split("?", 1)[0].split("#", 1)[0]

    if not url.startswith("/static/"):
        return None

    relative = url[len("/static/"):].lstrip("/\\")
    static_root = os.path.abspath(os.path.join(BASE_DIR, "static"))
    path = os.path.abspath(os.path.join(static_root, relative))

    # Evita que una ruta maliciosa salga de /static.
    if path != static_root and not path.startswith(static_root + os.sep):
        return None

    return path


def renderizar_dashboard_autonomo(datos_override=None):
    """
    Genera el mismo dashboard de /dashboard, pero convierte sus assets locales
    en recursos embebidos para que el archivo descargado pueda abrirse mediante
    file:/// sin depender de dashboard.css, dashboard.js ni imágenes externas.

    /dashboard sigue usando los assets separados de /static.
    /dashboard-file usa exclusivamente esta versión autónoma.
    """
    html = renderizar_dashboard_nueva_plantilla(datos_override)

    # 1) CSS: <link rel="stylesheet" href="/static/...css"> -> <style>...</style>
    patron_css = re.compile(
        r'<link\b(?=[^>]*\brel=["\']stylesheet["\'])(?=[^>]*\bhref=["\']([^"\']+)["\'])[^>]*>',
        re.IGNORECASE
    )

    def reemplazar_css(match):
        url = match.group(1)
        path = _dashboard_asset_path(url)

        if not path or not os.path.isfile(path):
            return match.group(0)

        try:
            css = open(path, "r", encoding="utf-8").read()
            return "<style>\n" + css + "\n</style>"
        except Exception:
            return match.group(0)

    html = patron_css.sub(reemplazar_css, html)

    # 2) JS externo: <script src="/static/...js"></script> -> <script>...</script>
    patron_js = re.compile(
        r'<script\b(?=[^>]*\bsrc=["\']([^"\']+)["\'])[^>]*>\s*</script>',
        re.IGNORECASE
    )

    def reemplazar_js(match):
        url = match.group(1)
        path = _dashboard_asset_path(url)

        if not path or not os.path.isfile(path):
            return match.group(0)

        try:
            js = open(path, "r", encoding="utf-8").read()
            # Evita cerrar el script si algún texto JS contiene esa secuencia.
            js = re.sub(r"</script", r"<\/script", js, flags=re.IGNORECASE)
            return "<script>\n" + js + "\n</script>"
        except Exception:
            return match.group(0)

    html = patron_js.sub(reemplazar_js, html)

    # 3) Imágenes de /static: src="/static/...png" -> data:image/...;base64,...
    patron_img = re.compile(
        r'(<img\b[^>]*?\bsrc=["\'])(/static/[^"\']+)(["\'][^>]*>)',
        re.IGNORECASE
    )

    def reemplazar_img(match):
        prefix, url, suffix = match.groups()
        path = _dashboard_asset_path(url)

        if not path or not os.path.isfile(path):
            return match.group(0)

        try:
            mime = mimetypes.guess_type(path)[0] or "application/octet-stream"
            with open(path, "rb") as fh:
                encoded = base64.b64encode(fh.read()).decode("ascii")
            return prefix + "data:" + mime + ";base64," + encoded + suffix
        except Exception:
            return match.group(0)

    html = patron_img.sub(reemplazar_img, html)

    return html



def construir_dashboard_html_autonomo():
    """
    Crea una copia del index.html con los datos actuales incrustados.

    El index.html original puede seguir usando fetch("./data.json?...").
    Para el archivo adjunto se intercepta únicamente la petición
    a data.json y se responde con los datos incrustados.

    Así el HTML puede abrirse desde file:/// sin CORS y sin necesitar
    un archivo data.json junto a él.
    """

    if not os.path.exists(
        DASHBOARD_TEMPLATE
    ):

        raise FileNotFoundError(
            "No existe index.html del dashboard: "
            + DASHBOARD_TEMPLATE
        )

    with open(
        DASHBOARD_TEMPLATE,
        "r",
        encoding="utf-8"
    ) as archivo:

        html = archivo.read()

    datos = cargar_dashboard_data()

    datos_json = json.dumps(
        datos,
        ensure_ascii=False
    )

    # Evita cerrar accidentalmente el bloque <script>.
    datos_json = datos_json.replace(
        "</",
        "<\\/"
    )

    script_embebido = """
<script>
(function () {
    const DASHBOARD_EMBEDDED_DATA = __DATOS__;

    window.__DASHBOARD_EMBEDDED_DATA__ =
        DASHBOARD_EMBEDDED_DATA;

    const originalFetch = window.fetch;

    window.fetch = function(resource, options) {

        let url = "";

        if (typeof resource === "string") {
            url = resource;
        } else if (
            resource &&
            typeof resource.url === "string"
        ) {
            url = resource.url;
        }

        if (
            url.includes("data.json")
        ) {

            return Promise.resolve(
                new Response(
                    JSON.stringify(
                        DASHBOARD_EMBEDDED_DATA
                    ),
                    {
                        status: 200,
                        headers: {
                            "Content-Type":
                                "application/json"
                        }
                    }
                )
            );
        }

        return originalFetch.call(
            window,
            resource,
            options
        );
    };
})();
</script>
""".replace(
        "__DATOS__",
        datos_json
    )

    # Debe cargarse antes que los scripts originales del dashboard.
    if "</head>" in html:

        html = html.replace(
            "</head>",
            script_embebido + "\\n</head>",
            1
        )

    elif "<body" in html:

        body_start = html.find("<body")
        posicion = html.find(
            ">",
            body_start
        )

        if posicion != -1:

            html = (
                html[:posicion + 1]
                + script_embebido
                + html[posicion + 1:]
            )

        else:

            html = (
                script_embebido
                + html
            )

    else:

        html = (
            script_embebido
            + html
        )

    return html


# ============================================================
# PROCESAR EXCEL
# ============================================================

@app.route(
    "/process-excel",
    methods=["POST"]
)
@require_api_key
def process_excel():

    if "file" not in request.files:

        return jsonify({
            "success": False,
            "error": "No se recibió ningún archivo"
        }), 400

    file = request.files["file"]

    if not file.filename:

        return jsonify({
            "success": False,
            "error": "El archivo no tiene nombre"
        }), 400

    if not file.filename.lower().endswith(
        ".xlsx"
    ):

        return jsonify({
            "success": False,
            "error": "El archivo debe ser .xlsx"
        }), 400

    if not os.path.exists(
        TEMPLATE_FILE
    ):

        return jsonify({
            "success": False,
            "error": "La plantilla no existe en el servidor",
            "template": TEMPLATE_FILE
        }), 500

    input_path = None
    output_path = None

    try:

        # ====================================================
        # GUARDAR ARCHIVO
        # ====================================================

        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as temp:

            file.save(
                temp.name
            )

            input_path = temp.name

        # ====================================================
        # ABRIR EXCEL DE ENTRADA
        # ====================================================

        workbook = load_workbook(
            input_path,
            data_only=True,
            read_only=True
        )

        # Moneda de respaldo del archivo completo.
        # Se obtiene del texto real / país / bancos antes de procesar hojas.
        moneda_archivo = detectar_moneda_contexto_archivo(
            workbook,
            file.filename
        )

        print(
            "Moneda contextual detectada:",
            moneda_archivo
        )

        # ====================================================
        # IDENTIFICACION BANCARIA
        # ====================================================
        # Camino principal: texto (rápido). OCR se activa SOLO para hojas donde
        # el banco no pudo identificarse por nombre de hoja/celdas. Además está
        # limitado a pocas imágenes y a un presupuesto total de tiempo para no
        # comprometer la respuesta de /process-excel dentro de Make/Render.
        bancos_ocr_por_hoja = {}

        # Si el nombre del archivo ya identifica claramente el banco, NO ejecutar
        # OCR. Esto mantiene rápido /process-excel y evita falsos positivos.
        banco_archivo = detectar_banco_por_nombre_archivo(
            file.filename
        )

        if banco_archivo:
            print(
                "Banco detectado por nombre de archivo:",
                file.filename,
                "=>",
                banco_archivo
            )
        else:
            try:
                hojas_para_ocr = _hojas_que_requieren_ocr(
                    workbook
                )

                if hojas_para_ocr:
                    print(
                        "Hojas candidatas a OCR bancario:",
                        hojas_para_ocr
                    )

                    bancos_ocr_por_hoja = detectar_bancos_desde_imagenes_excel(
                        input_path,
                        hojas_objetivo=hojas_para_ocr,
                        max_images_per_sheet=OCR_MAX_IMAGES_PER_SHEET,
                        max_seconds=OCR_MAX_SECONDS_PER_FILE
                    )

            except Exception as ocr_error:
                # El OCR es solo respaldo: cualquier problema debe dejar continuar
                # el procesamiento normal del archivo.
                print(
                    "Advertencia: OCR bancario omitido:",
                    str(ocr_error)
                )
                bancos_ocr_por_hoja = {}

        all_transactions = []
        processed_sheets = []

        # ====================================================
        # PROCESAR TODAS LAS HOJAS
        # ====================================================

        for worksheet in workbook.worksheets:

            try:

                transactions = procesar_hoja(
                    worksheet,
                    moneda_respaldo=moneda_archivo,
                    nombre_archivo=file.filename,
                    banco_respaldo_ocr=bancos_ocr_por_hoja.get(
                        worksheet.title,
                        ""
                    )
                )

                if transactions:

                    all_transactions.extend(
                        transactions
                    )

                    processed_sheets.append({
                        "name": worksheet.title,
                        "transactions": len(
                            transactions
                        ),
                        "currency": (
                            transactions[0].get("moneda", "N/D")
                            if transactions
                            else "N/D"
                        ),
                        "country": (
                            transactions[0].get("pais", "NO_IDENTIFICADO")
                            if transactions
                            else "NO_IDENTIFICADO"
                        ),
                        "bank": (
                            transactions[0].get("banco", "N/D")
                            if transactions
                            else "N/D"
                        ),
                        "bank_source": (
                            "ocr_image"
                            if worksheet.title in bancos_ocr_por_hoja
                            else "text"
                        )
                    })

            except Exception as sheet_error:

                # Una hoja problemática no debe tumbar
                # todo el procesamiento.

                processed_sheets.append({
                    "name": worksheet.title,
                    "transactions": 0,
                    "error": str(sheet_error)
                })

        workbook.close()

        # ====================================================
        # VALIDAR
        # ====================================================

        if not all_transactions:

            return jsonify({
                "success": False,
                "error": (
                    "No se encontraron movimientos "
                    "bancarios compatibles"
                ),
                "processed_sheets": processed_sheets
            }), 400

        # ====================================================
        # MODO LOTE
        # ====================================================

        batch_id = _sanitizar_batch_id(
            request.form.get(
                "batch_id",
                ""
            )
        )

        file_id = str(
            request.form.get(
                "file_id",
                ""
            )
            or ""
        ).strip()

        if batch_id:

            state = _agregar_archivo_a_lote(
                batch_id=batch_id,
                file_id=file_id,
                file_name=file.filename,
                transactions=all_transactions,
                processed_sheets=processed_sheets,
                currency=moneda_archivo
            )

            if (
                input_path
                and os.path.exists(
                    input_path
                )
            ):
                try:
                    os.remove(
                        input_path
                    )
                    input_path = None
                except Exception:
                    pass

            files_count = len(
                state.get(
                    "files",
                    {}
                )
            )

            total_transactions = sum(
                len(
                    item.get(
                        "transactions",
                        []
                    )
                    or []
                )
                for item
                in state.get(
                    "files",
                    {}
                ).values()
            )

            return jsonify({
                "success": True,
                "mode": "batch",
                "batch_id": batch_id,
                "file_id": file_id,
                "file_name": file.filename,
                "file_transactions": len(
                    all_transactions
                ),
                "files_in_batch": files_count,
                "transactions_in_batch": total_transactions,
                "currency": moneda_archivo,
                "processed_sheets": processed_sheets,
                "message": (
                    "Archivo agregado al lote. "
                    "El dashboard se generará en /finalize-batch."
                )
            })

        # ====================================================
        # MODO INDIVIDUAL / COMPATIBILIDAD
        # ====================================================

        # ====================================================
        # ARCHIVO FINAL TEMPORAL
        # ====================================================

        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as temp:

            output_path = temp.name

        # ====================================================
        # CREAR RESULTADO
        # ====================================================

        insertar_en_plantilla(
            all_transactions,
            output_path
        )


        # ====================================================
        # ACTUALIZAR DASHBOARD HTML
        # ====================================================

        dashboard_payload = construir_dashboard_data(
            all_transactions,
            file.filename,
            moneda_preferida=moneda_archivo
        )

        guardar_dashboard_data(
            dashboard_payload
        )

        # ====================================================
        # ELIMINAR INPUT
        # ====================================================

        if (
            input_path
            and os.path.exists(input_path)
        ):

            try:

                os.remove(
                    input_path
                )

                input_path = None

            except Exception:

                pass

        # ====================================================
        # LIMPIAR OUTPUT DESPUÉS DE ENVIAR
        # ====================================================

        @after_this_request
        def cleanup(response):

            try:

                if (
                    output_path
                    and os.path.exists(
                        output_path
                    )
                ):

                    os.remove(
                        output_path
                    )

            except Exception:

                pass

            return response

        # ====================================================
        # DEVOLVER EXCEL
        # ====================================================

        response = send_file(
            output_path,
            as_attachment=True,
            download_name=OUTPUT_FILENAME,
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
        response.headers["X-Parser-Version"] = PARSER_VERSION
        response.headers["X-Parser-Transactions"] = str(len(all_transactions))
        response.headers["X-Parser-Sheets"] = "; ".join(
            f"{item.get('name')}={item.get('transactions', 0)}"
            for item in processed_sheets
        )[:1500]

        response.headers[
            "X-Dashboard-URL"
        ] = (
            request.host_url.rstrip("/")
            + "/dashboard"
        )


        response.headers[
            "X-Dashboard-File-URL"
        ] = (
            request.host_url.rstrip("/")
            + "/dashboard-file"
        )

        response.headers[
            "X-Dashboard-Attachment-Version"
        ] = DASHBOARD_ATTACHMENT_VERSION

        response.headers[
            "X-Dashboard-Banks"
        ] = str(
            dashboard_payload.get(
                "meta", {}
            ).get(
                "banks", 0
            )
        )

        response.headers[
            "X-Dashboard-Balance"
        ] = str(
            dashboard_payload.get(
                "totals", {}
            ).get(
                "final", 0
            )
        )

        response.headers["X-Dashboard-Currency"] = str(
            dashboard_payload.get(
                "meta", {}
            ).get(
                "currency", "N/D"
            )
        )

        response.headers["X-Source-Currency"] = str(
            moneda_archivo
        )

        response.headers["X-App-Build"] = APP_BUILD
        return response

    except Exception as e:

        return jsonify({
            "success": False,
            "error": str(e),
            "error_type": type(e).__name__,
            "app_build": APP_BUILD,
            "parser_version": PARSER_VERSION if "PARSER_VERSION" in globals() else "legacy"
        }), 500

    finally:

        if (
            input_path
            and os.path.exists(
                input_path
            )
        ):

            try:

                os.remove(
                    input_path
                )

            except Exception:

                pass



# ============================================================
# ENDPOINTS DE LOTES
# ============================================================

@app.route(
    "/finalize-batch",
    methods=["POST"]
)
@require_api_key
def finalize_batch():

    payload = request.get_json(
        silent=True
    ) or {}

    batch_id = _sanitizar_batch_id(
        payload.get(
            "batch_id"
        )
        or request.form.get(
            "batch_id",
            ""
        )
        or request.args.get(
            "batch_id",
            ""
        )
    )

    if not batch_id:
        return jsonify({
            "success": False,
            "error": "batch_id requerido"
        }), 400

    try:
        state = _cargar_estado_lote(
            batch_id,
            create=False
        )

        if not state:
            return jsonify({
                "success": False,
                "error": "No existe ese lote",
                "batch_id": batch_id
            }), 404

        files = state.get(
            "files",
            {}
        )

        if not files:
            return jsonify({
                "success": False,
                "error": "El lote no contiene archivos",
                "batch_id": batch_id
            }), 400

        all_transactions = _transacciones_del_lote(
            state
        )

        if not all_transactions:
            return jsonify({
                "success": False,
                "error": "El lote no contiene movimientos bancarios compatibles",
                "batch_id": batch_id
            }), 400

        result_path = _batch_result_path(
            batch_id
        )

        insertar_en_plantilla(
            all_transactions,
            result_path
        )

        batch_file_name = (
            "Lote "
            + batch_id
            + " - "
            + str(
                len(
                    files
                )
            )
            + " archivos"
        )

        dashboard_payload = construir_dashboard_data(
            all_transactions,
            batch_file_name,
            moneda_preferida=_moneda_preferida_lote(
                state
            )
        )

        # Mantener también el dashboard global para compatibilidad.
        guardar_dashboard_data(
            dashboard_payload
        )

        state["dashboard_payload"] = dashboard_payload
        state["result_path"] = result_path
        state["finalized"] = True
        state["finalized_at"] = datetime.now().isoformat(
            timespec="seconds"
        )

        _guardar_estado_lote(
            batch_id,
            state
        )

        base_url = request.host_url.rstrip(
            "/"
        )

        dashboard_url = (
            base_url
            + "/dashboard?batch_id="
            + batch_id
        )

        dashboard_file_url = (
            base_url
            + "/dashboard-file?batch_id="
            + batch_id
        )

        excel_url = (
            base_url
            + "/result-excel?batch_id="
            + batch_id
        )

        countries = dashboard_payload.get(
            "meta",
            {}
        ).get(
            "countries",
            []
        )

        return jsonify({
            "success": True,
            "batch_id": batch_id,
            "files_processed": len(
                files
            ),
            "transactions": len(
                all_transactions
            ),
            "banks": dashboard_payload.get(
                "meta",
                {}
            ).get(
                "banks",
                0
            ),
            "accounts": dashboard_payload.get(
                "meta",
                {}
            ).get(
                "accounts",
                0
            ),
            "countries": countries,
            "currency": dashboard_payload.get(
                "meta",
                {}
            ).get(
                "currency",
                "N/D"
            ),
            "dashboard_url": dashboard_url,
            "dashboard_file_url": dashboard_file_url,
            "excel_url": excel_url,
            "message": (
                "Lote finalizado. Se generó un único Excel "
                "y un único dashboard consolidado."
            )
        })

    except Exception as error:
        return jsonify({
            "success": False,
            "error": str(
                error
            ),
            "error_type": type(
                error
            ).__name__,
            "batch_id": batch_id,
            "app_build": APP_BUILD
        }), 500


@app.route(
    "/result-excel",
    methods=["GET"]
)
@require_api_key
def result_excel_batch():

    batch_id = _sanitizar_batch_id(
        request.args.get(
            "batch_id",
            ""
        )
    )

    if not batch_id:
        return jsonify({
            "success": False,
            "error": "batch_id requerido"
        }), 400

    state = _cargar_estado_lote(
        batch_id,
        create=False
    )

    if not state:
        return jsonify({
            "success": False,
            "error": "Lote no encontrado",
            "batch_id": batch_id
        }), 404

    result_path = state.get(
        "result_path"
    ) or _batch_result_path(
        batch_id
    )

    if (
        not state.get(
            "finalized"
        )
        or not os.path.exists(
            result_path
        )
    ):
        return jsonify({
            "success": False,
            "error": "El lote todavía no ha sido finalizado",
            "batch_id": batch_id
        }), 409

    return send_file(
        result_path,
        as_attachment=True,
        download_name=OUTPUT_FILENAME,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


@app.route(
    "/batch-status",
    methods=["GET"]
)
@require_api_key
def batch_status():

    batch_id = _sanitizar_batch_id(
        request.args.get(
            "batch_id",
            ""
        )
    )

    if not batch_id:
        return jsonify({
            "success": False,
            "error": "batch_id requerido"
        }), 400

    state = _cargar_estado_lote(
        batch_id,
        create=False
    )

    if not state:
        return jsonify({
            "success": False,
            "error": "Lote no encontrado",
            "batch_id": batch_id
        }), 404

    files = state.get(
        "files",
        {}
    )

    return jsonify({
        "success": True,
        "batch_id": batch_id,
        "files_in_batch": len(
            files
        ),
        "transactions_in_batch": sum(
            len(
                item.get(
                    "transactions",
                    []
                )
                or []
            )
            for item
            in files.values()
        ),
        "finalized": bool(
            state.get(
                "finalized"
            )
        ),
        "files": [
            {
                "file_id": item.get(
                    "file_id",
                    ""
                ),
                "file_name": item.get(
                    "file_name",
                    ""
                ),
                "currency": item.get(
                    "currency",
                    "N/D"
                )
            }
            for item
            in files.values()
        ]
    })


# ============================================================
# SERVIDOR
# ============================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=int(
            os.environ.get(
                "PORT",
                5000
            )
        ),
        debug=False
    )   